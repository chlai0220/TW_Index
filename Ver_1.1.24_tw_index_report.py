# -*- conding: utf-8 -*-
####################################################################################
# Python Version : 3.4.3
# File Name      : 
# Author         : C.H. Lai
# Purpose        : 
####################################################################################
# History: 
#  Ver        Date               Descripition
# -------  ------------  ----------------------------
#  1.0.0     2017/10/12   * New Issue
#  1.1.0     2017/11/01   * Init Release
#  1.1.1     2017/11/02   * Fix Bub For [Fun]exe_Time
#                         * Modify Directory Rule For Baclup
#  1.1.2     2017/11/03   * Add Update Checking Function
#  1.1.3     2017/11/06   * Add Function For TW Option
#  1.1.4     2017/11/07   * Fix Bug For TXO
#  1.1.5     2017/11/10   * Fix Bug For TXO - RE Function
#  1.1.6     2017/11/13   * Add TXO Report
#                         * Modify GMail Sending Function
#  1.1.6A    2017/11/19   * GMail Check Function For Auto-Sending
#  1.1.6B    2017/11/22   * Fix Error For TXO Excel File
#  1.1.7     2017/11/27   * Add TW/TX Daily Report For Excel File
#  1.1.8     2017/12/07   * Add TXO At-The-Money Price Daily Report
#  1.1.9     2018/03/21   * Add Month TXO At-The-Money Price Daily Report
#  1.1.10    2018/05/05   * Modify TXO Weekly-OP Report
#  1.1.11    2018/06/20   * Modify Excel Format For TW Daily Report
#  1.1.12    2018/07/30   * Remove Function - Fetch_Error
#  1.1.13    2018/07/31   * Add Function For TW Future OI Function
#  1.1.14    2018/10/02   * Modify TX/TXO/TXF For New Website Url
#  1.1.15    2018/10/17   * Modify TX For Website Fetch Format
#  1.1.16    2018/10/18   * Modify TX For Website Fetch Format
#  1.1.16A   2018/10/19   * Modify TX For RE Fetch Format
#  1.1.16B   2018/10/21   * Modify TX For RE Fetch Format
#  1.1.16C   2018/10/23   * Modify TX For RE Fetch Format
#  1.1.17    2018/11/21   * Modify TX For RE Fetch Format
#  1.1.18    2018/12/28   * Add Small-TXF Function
#  1.1.19    2019/01/03   * Add TXO Option Function
#  1.1.20    2019/01/07   * Add Weekday Check Function For 2019 New Rule Of TW Stcok
#  1.1.21    2019/01/16   * Modify TX For RE Fetch Format
#  1.1.22    2019/01/21   * Modify Check Rule
#  1.1.23    2019/03/06   * 新增外資選擇權成本價計算
#  1.1.24    2019/06/18   * Remove Report Function For Uploading Gitgub
####################################################################################
#+--------------------+
#|       MODULE       |     
#+--------------------+
#==================================
import csv
import sys , os
import sqlite3
import time
import re
import zipfile
import datetime
import traceback

# <<< Url Function >>>
import urllib.request
import socket
import requests

#<<< OpenpyXL >>>
from openpyxl import Workbook
from datetime import date
from openpyxl.chart import (
    LineChart,
    BarChart,
    StockChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.styles import colors
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

#<<< SMTP >>>
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
#==================================

#*******************************************#
# FUNCTION 
#*******************************************#
##############################################################################
 
##############################################################################
def Read_DB(DB_Type , Table_No):
    if DB_Type.upper()=="TW":
        my_db = DB_Path()[0]    #["twse.db", "tx.db"]
        #....................................................
        # Table #0 : working_day
        # Table #1 : daily
        # Table #2 : weekly
        # Table #3 : monthly
        #....................................................
        if Table_No<=-1 or Table_No>3:
            print("TW DB Table List Error!!!Please Double Confirm")
            return(None)
        else:
            myTable = {0:"working_day" ,  1:"daily" , 2:"weekly" , 3:"monthly"}
            
    elif DB_Type.upper()=="TX":
        #...........................................................................
        # Table #0 : source
        # Table #1 : working_day
        # Table #2 : contract_list  #契約列表
        # Table #3 : daily
        # Table #4 : weekly
        # Table #5 : monthly
        #...........................................................................
        my_db = DB_Path()[1]    #["twse.db", "tx.db"]
        
        if Table_No<=-1 or Table_No>6:
            print("TX DB Table List Error!!!Please Double Confirm")
            return(None)
        else:
            myTable = {0 : "source" , 1 : "working_day" ,  2 : "contract_list" , 3 : "daily" , 4 : "weekly" , 5 : "monthly"}

    elif DB_Type.upper()=="TXO":
        #...........................................................................
        # Table #0 : settlement_list
        # Table #1 : working_day
        #...........................................................................
        my_db = DB_Path()[2]    #["twse.db", "tx.db" , "txo.db"]
        
        if Table_No<=-1 or Table_No>1:
            print("TXO DB Table List Error!!!Please Double Confirm")
            return(None)
            
        else:
            myTable = {0 : "settlement_list" , 1 : "working_day"}
            
    #TXF
    elif DB_Type.upper()=="TXF":
        #...........................................................................
        # Table #0 : working_day
        # Table #1 : ii1
        # Table #2 : ii2
        # Table #3 : ii3
        #...........................................................................
        my_db = DB_Path()[3]    #["twse.db", "tx.db" , "txo.db" , "txf.db"]
        
        if Table_No<=-1 or Table_No>3:
            print("TXO DB Table List Error!!!Please Double Confirm")
            return(None)
            
        else:
            myTable = {0 : "working_day" , 1 : "ii1" ,  2 : " ii2" , 3 : "ii3"}

    elif DB_Type.upper()=="STXF":
        #...........................................................................
        # Table #0 : working_day
        # Table #1 : ii1
        # Table #2 : ii2
        # Table #3 : ii3
        #...........................................................................
        my_db = DB_Path()[4]    #["twse.db", "tx.db" , "txo.db" , "txf.db" , "stxf.db"]
        
        if Table_No<=-1 or Table_No>3:
            print("TXO DB Table List Error!!!Please Double Confirm")
            return(None)
            
        else:
            myTable = {0 : "working_day" , 1 : "ii1" ,  2 : " ii2" , 3 : "ii3"}
            
    elif DB_Type.upper()=="TXOP":
        #Ver 1.1.19
        #...........................................................................
        # Table #0 : working_day
        # Table #1 : ii1
        # Table #2 : ii2
        # Table #3 : ii3
        #...........................................................................
        my_db = DB_Path()[5]    #["twse.db", "tx.db" , "txo.db" , "txf.db" , "stxf.db" , "txop.db"]
        
        if Table_No<=-1 or Table_No>3:
            print("TXO DB Table List Error!!!Please Double Confirm")
            return(None)
            
        else:
            myTable = {0 : "working_day" , 1 : "ii1" ,  2 : " ii2" , 3 : "ii3"}

    else:
        print("DB Name Error!!!Please Double Confirm")
        return(None)
    #DB SETUP
    my_table = myTable[Table_No]
    
    #
    #print(my_db , my_table)
    conn = sqlite3.connect(my_db)
    cur = conn.cursor()
    cmd = "SELECT * FROM " + my_table
    cur.execute(cmd)
    querty = cur.fetchall()
    conn.close()
    return(querty)
    #---------- DONE ----------

##############################################################################

def DB_Insert(DB_Type , Table_No , Content):
    #...........................................................................
    if DB_Type.upper()=="TW":
        my_db = DB_Path()[0]    #["twse.db", "tx.db"]
        
        if Table_No==0:
            table0 = "working_day"
            cmd = "INSERT OR REPLACE INTO " + table0+ " VALUES(?,?,?,?)"
                
        elif Table_No==1:
            table1 = "daily"
            cmd = "INSERT OR REPLACE INTO " + table1+ " VALUES(?,?,?,?,?,?,?,?,?,?,?)"
                
        elif Table_No==2:
            table2 = "weekly"
            cmd = "INSERT OR REPLACE INTO " + table2+ " VALUES(?,?,?,?,?,?,?,?)"
                
        elif Table_No==3:
            table3 = "monthly"
            cmd = "INSERT OR REPLACE INTO " + table3+ " VALUES(?,?,?,?,?,?,?,?)"
                
        else:
            print("TW DB Table No Error!")
            print("Please Check Table No Again")
            print("Stop Procedure!")
            input("Please Enter To Leave...")
            sys.exit(0)
    #. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 
    elif DB_Type.upper()=="TX":
        my_db = DB_Path()[1]    #["twse.db", "tx.db"]
        
        if Table_No==0:
            table0 = "source"
            cmd = "INSERT OR REPLACE INTO " + table0 + " VALUES(?,?,?,?,?,?,?,?,?,?)"
                
        elif Table_No==1:
            table1 = "working_day"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?,?,?,?)"
                
        elif Table_No==2:
            table2 = "contract_list"
            cmd = "INSERT OR REPLACE INTO " + table2 + " VALUES(?)"
                
        elif Table_No==3:
            table3 = "daily"
            cmd = "INSERT OR REPLACE INTO " + table3 + " VALUES(?,?,?,?,?,?,?,?,?,?)"
            
        elif Table_No==4:
            table4 = "weekly"
            cmd = "INSERT OR REPLACE INTO " + table4 + " VALUES(?,?,?,?,?,?,?)"
            
        elif Table_No==5:
            table5 = "monthly"
            cmd = "INSERT OR REPLACE INTO " + table5 + " VALUES(?,?,?,?,?,?,?)"
                
        else:
            print("TW DB Table No Error!")
            print("Please Check Table No Again")
            print("Stop Procedure!")
            input("Please Enter To Leave...")
            sys.exit(0)
    #. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 
    elif DB_Type.upper()=="TXO":
        my_db = DB_Path()[2]    #["twse.db", "tx.db" , "txo.db"]
        
        if Table_No==0:
            table0 = "settlement_list"
            cmd = "INSERT OR REPLACE INTO " + table0 + " VALUES(?,?,?)"
            
        elif Table_No==1:
            table1 = "working_day"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?)"
            
        else:
            print("TWO DB Table No Error!")
            print("Please Check Table No Again")
            print("Stop Procedure!")
            input("Please Enter To Leave...")
            sys.exit(0)
    #. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 
    elif DB_Type.upper()=="TXF":
        my_db = DB_Path()[3]    #["twse.db", "tx.db" , "txo.db" , "txf.db"]

        if Table_No==0:
            table0 = "working_day"
            cmd = "INSERT OR REPLACE INTO " + table0 + " VALUES(?)"
            
        elif Table_No==1:
            table1 = "ii1"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?,?,?,?,?,?,?)"
            
        elif Table_No==2:
            table2 = "ii2"
            cmd = "INSERT OR REPLACE INTO " + table2 + " VALUES(?,?,?,?,?,?,?)"
            
        elif Table_No==3:
            table3 = "ii3"
            cmd = "INSERT OR REPLACE INTO " + table3 + " VALUES(?,?,?,?,?,?,?)"
            
        else:
            print("TXF DB Table No Error!")
            print("Please Check Table No Again")
            print("Stop Procedure!")
            input("Please Enter To Leave...")
            sys.exit(0)
    #. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 
    elif DB_Type.upper()=="STXF":
        my_db = DB_Path()[4]    #["twse.db", "tx.db" , "txo.db" , "txf.db"]

        if Table_No==0:
            table0 = "working_day"
            cmd = "INSERT OR REPLACE INTO " + table0 + " VALUES(?)"
            
        elif Table_No==1:
            table1 = "ii1"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?,?,?,?,?,?,?)"
            
        elif Table_No==2:
            table2 = "ii2"
            cmd = "INSERT OR REPLACE INTO " + table2 + " VALUES(?,?,?,?,?,?,?)"
            
        elif Table_No==3:
            table3 = "ii3"
            cmd = "INSERT OR REPLACE INTO " + table3 + " VALUES(?,?,?,?,?,?,?)"
            
        else:
            print("STXF DB Table No Error!")
            print("Please Check Table No Again")
            print("Stop Procedure!")
            input("Please Enter To Leave...")
            sys.exit(0)
    #. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 
    elif DB_Type.upper()=="TXOP":

        my_db = DB_Path()[5]    #["twse.db", "tx.db" , "txo.db" , "txf.db" , "txop.db"]
        
        if Table_No==0:
            table1 = "working_day"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?)"
            
        elif Table_No==1:
            table1 = "ii1"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)"
            
        elif Table_No==2:
            table1 = "ii2"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)"
            
        elif Table_No==3:
            table1 = "ii3"
            cmd = "INSERT OR REPLACE INTO " + table1 + " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)"
        else:
            print("TWOP DB Table No Error!")
            print("Please Check Table No Again")
            print("Stop Procedure!")
            input("Please Enter To Leave...")
            sys.exit(0)    #. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . 
    else:
        print("TW DB Type Error!")
        print("Please Check Table No Again")
        print("Stop Procedure!")
        input("Please Enter To Leave...")
        sys.exit(0)    
    #
    conn = sqlite3.connect(my_db)
    conn.executemany(cmd , Content)
    conn.commit()
    conn.close()
    #...........................................................................
    
##############################################################################
class TW_DB_Create:
    def __init__(self):
        self.DB_Create()

    def DB_Create(self):
        my_db = DB_Path()[0]    #["twse.db", "tx.db"]
        #...........................................................................
        # Raw Data DB - Table Lists
        # Table #0 : working_day
        # Table #1 : daily
        # Table #2 : weekly
        # Table #3 : monthly
        #...........................................................................
        conn = sqlite3.connect(my_db)
        #...........................................................................
        # <<<<< Table : working day >>>>>
        #...........................................................................
        table0 = "working_day"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table0 + """(
               Date    TEXT    NOT NULL,
               Wkno  TEXT    NOT NULL,
               Year     TEXT    NOT NULL,
               Mon     TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table0  + "(Date)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        # <<<<< Table : daily >>>>>
        # 日期 週別 年份 月份 開盤指數 最高指數 最低指數  收盤指數   成交金額    成交股數 成交筆數
        # 成交金額 = Turnover
        # 成交量 = Volume
        # 成交筆數 = Transactions
        #...........................................................................
        table1 = "daily"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table1 + """(
               Date    TEXT    NOT NULL,
               Wkno  TEXT    NOT NULL,
               Year     TEXT    NOT NULL,
               Mon     TEXT    NOT NULL,
               Open   REAL    NOT NULL,
               High     REAL    NOT NULL,
               Low      REAL    NOT NULL,
               Close   REAL    NOT NULL,
               Turnover     INT     NOT NULL,
               Volume      INT     NOT NULL,
               Trans      INT     NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table1  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : weekly >>>>>
        # 週別 開盤指數 最高指數 最低指數  收盤指數   成交金額    成交股數 成交筆數
        # 成交金額 = Turnover
        # 成交量 = Volume
        # 成交筆數 = Transactions
        #...........................................................................
        table2 = "weekly"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table2 + """(
               Wkno  TEXT    NOT NULL,
               Open   REAL    NOT NULL,
               High     REAL    NOT NULL,
               Low      REAL    NOT NULL,
               Close   REAL    NOT NULL,
               Turnover     INT     NOT NULL,
               Volume      INT     NOT NULL,
               Trans      INT     NOT NULL,
               UNIQUE(Wkno))"""
        conn.execute(cmd)
        
        #Index - Wkno
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Wkno ON " + table2  + "(Wkno)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        # <<<<< Table : monthly >>>>>
        # 年份 開盤指數 最高指數 最低指數  收盤指數   成交金額    成交股數 成交筆數
        # 成交金額 = Turnover
        # 成交量 = Volume
        # 成交筆數 = Transactions
        #...........................................................................
        table3 = "monthly"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table3 + """(
               Mon     TEXT    NOT NULL,
               Open   REAL    NOT NULL,
               High     REAL    NOT NULL,
               Low      REAL    NOT NULL,
               Close   REAL    NOT NULL,
               Turnover     REAL     NOT NULL,
               Volume      INT     NOT NULL,
               Trans      INT     NOT NULL,
               UNIQUE(Mon))"""
        conn.execute(cmd)
        
        #Index - Mon
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Mon ON " + table3  + "(Mon)"
        conn.execute(cmd)
        #...........................................................................
        conn.close()
        #---------- DONE ----------

##############################################################################
        
class TW_Index_By_Daily:
    def __init__(self):
        self.Data_Fetch()

    def Data_Fetch(self):
        #...........................................................................
        data_num = len(Read_DB("TW",0))
        if data_num==0:
            print("Start Init Set-up For TWSE")
            (tw_index_files , tw_volume_files) = self.Init_Update()
        else:
            print("Start Normal Set-up For TWSE")
            (tw_index_files , tw_volume_files) = self.Nor_Update()
            
        self.Csv_Reader(tw_index_files , tw_volume_files)
        #---------- DONE ----------
        #...........................................................................

    def Init_Update(self):
        #...........................................................................
        #加權指數     ※ 本資料自民國88年01月05日開始提供
        #成交量資料 ※ 本資訊自民國79年1月4日起開始提供
        nyear = int(time.strftime("%Y", time.localtime()))
        nmon = int(time.strftime("%m", time.localtime()))
        #88+1911 = 1999
        tw_index_files = []
        tw_volume_files = []
        for yy in range(1999 , nyear):
            for mm in range(1,13):
                (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)                
                
        for yy in range(nyear , nyear+1):
            for mm in range(1,nmon+1):
                (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)
        return(tw_index_files , tw_volume_files)
        #---------- DONE ----------
        #...........................................................................
                
    def Nor_Update(self):
        #...........................................................................
        now_date = time.strftime("%Y-%m-%d" , time.localtime())
        nyear = int(time.strftime("%Y", time.localtime()))
        nmon = int(time.strftime("%m", time.localtime()))
        tw_index_files = []
        tw_volume_files = []

        (db_date , db_wkno , db_year , db_mon) = Read_DB("TW" , 0)[-1] #('2017-10-16', '2017WK42', '2017', '201710')

        now_year = time.strftime("%Y" , time.localtime())
        now_mon = time.strftime("%Y%m" , time.localtime())

        #if not db_date==now_date:
        if True:
            #*** Rule1 ***
            if int(now_year)-int(db_year)==0:
                #當年度
                if db_mon[-2:]==now_mon[-2:]:
                    for yy in range(nyear , nyear+1):
                        for mm in range(nmon,nmon+1):
                            (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)
                else:
                    for yy in range(nyear , nyear+1):
                        for mm in range(int(db_mon[-2:]),nmon+1):
                            (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)
                        
            #*** Rule1 ***
            elif nyear-int(db_year)==1:
                #相差1年                
                for yy in range(nyear-1 , nyear):
                    for mm in range(int(db_mon[-2:]),13):
                        (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)
                        
                for yy in range(nyear , nyear+1):
                    for mm in range(1,nmon+1):
                        (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)
                        
            #*** Rule3 ***
            elif nyear-int(db_year)>=2:
                #相差2年以上
                for yy in range(int(db_year) , int(db_year)+1):
                    for mm in range(int(db_mon[-2:]),13):
                        (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)

                for yy in range(int(db_year)+1 , nyear):
                    for mm in range(1,13):
                        (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)

                for yy in range(nyear , nyear+1):
                    for mm in range(1,nmon+1):
                        (tw_index_files ,tw_volume_files) = self.Csv_Download(yy , mm , tw_index_files ,tw_volume_files)
            
        return(tw_index_files , tw_volume_files)
        #---------- DONE ----------
        #...........................................................................

    def Csv_Download(self , year , mon , tw_index_files , tw_volume_files):
        #...........................................................................
        #ch_dirs = ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
        ch_dir1 = os.path.join(Dir_Chk()[-1]  , str(year))
        ch_dir2 = os.path.join(ch_dir1  , str(year)+ str(mon).zfill(2))
        
        if not os.path.exists(ch_dir1):
            os.mkdir(ch_dir1)
            
        if not os.path.exists(ch_dir2):
            os.mkdir(ch_dir2)
        #...........................................................................
        #加權指數
        myUrl1 = 'http://www.twse.com.tw/indicesReport/MI_5MINS_HIST?response=csv&date=' + str(year) + str(mon).zfill(2) + '01'
            
        file_name1 = os.path.join(ch_dir2 , (str(year) + str(mon).zfill(2) + "-TWSE1.csv"))

        #暫停5S
        time.sleep(3)
        
        urllib.request.urlretrieve(myUrl1, file_name1)
        
        tw_index_files.append(file_name1)
        #...........................................................................
        myUrl2 = 'http://www.twse.com.tw/exchangeReport/FMTQIK?response=csv&date=' + str(year) + str(mon).zfill(2) + '01'

        file_name2 = os.path.join(ch_dir2 , (str(year) + str(mon).zfill(2) + "-TWSE2.csv"))

        #暫停5S
        time.sleep(3)
        
        urllib.request.urlretrieve(myUrl2, file_name2)
        
        tw_volume_files.append(file_name2)
        #...........................................................................
        #print(file_name1 , file_name2)
        return(tw_index_files , tw_volume_files)
        #---------- DONE ----------
        #...........................................................................

    def Csv_Date_Handle(self , ch_date):
        #...........................................................................
        chDates = ch_date.split("/")
        
        ceDate = str(int(chDates[0]) +1911) + "-" + chDates[1] + "-" + chDates[2]
        
        t = time.strptime(str(int(chDates[0]) +1911) + "/" + chDates[1] + "/" + chDates[2], "%Y/%m/%d")
        wkno = time.strftime("%Y%U" , t)
            
        if int(wkno)==0:
            t = time.strptime(str(int(chDates[0]) +1911-1)+"/12/31" , "%Y/%m/%d")
            wkno = time.strftime("%YWK%U" , t)
        else:
            wkno = time.strftime("%YWK%U" , t)
        return(ceDate , wkno , time.strftime("%Y" , t) , time.strftime("%Y%m" , t) )
        #---------- DONE ----------
        #...........................................................................

    def Csv_Reader(self , tw_index_files , tw_volume_files):
        #...........................................................................
        temp1 = []
        temp2 = []
        content = []
        date_content = []
        #...........................................................................
        csv_num1 = len(tw_index_files)
        csv_num2 = len(tw_volume_files)
        if csv_num1==csv_num2:
            for i in range(csv_num1):
                file1 = tw_index_files[i]
                file2 = tw_volume_files[i]
                
                fp1 = open(file1 , 'r')
                for i , row1 in enumerate(csv.reader(fp1)):
                    if len(row1)>0 and (len(row1[0])==8 or len(row1[0])==9):
                        #print(row1)
                        temp1.append(row1)
                fp1.close()
                
                fp2 = open(file2 , 'r')
                for i , row1 in enumerate(csv.reader(fp2)):
                    #if len(row1)>0:
                    if len(row1)>0 and (len(row1[0])==8 or len(row1[0])==9):
                        #print(row1)
                        temp2.append(row1)
                fp2.close()
        #...........................................................................
        len1 = len(temp1)
        len2 = len(temp2)
        #print(len1 , len2)
        if len1==len2:
            for i in range(len1):
                row_temp = []
                
                (ceDate , wkno , year , mon) = self.Csv_Date_Handle(temp1[i][0])
                #print(ceDate , wkno , year , mon)
                row_temp.append(ceDate)
                row_temp.append(wkno)
                row_temp.append(year)
                row_temp.append(mon)
                for j in range(1,5):
                    row_temp.append(temp1[i][j])
                row_temp.append(float(int(temp2[i][2].replace("," , ""))/100000000))
                row_temp.append(int(temp2[i][1].replace("," , "")))
                row_temp.append(int(temp2[i][3].replace("," , "")))

                date_content.append(row_temp[:4])
                content.append(row_temp)
                        
        #return(date_content , content)
        DB_Insert("TW" , 0 , date_content)
        DB_Insert("TW" , 1 , content)

        #---------- DONE ----------
        #...........................................................................
        


##############################################################################

class TW_Index_By_Weekly:
    def __init__(self):
        self.DB_data_sum()

    def DB_data_sum(self):
        #...........................................................................
        wkno_list = sorted(set([item[1] for item in Read_DB("TW" , 0)]))
        
        if len(Read_DB("TW" , 2))>0:
            if len(wkno_list)>6:
                wkno_list = wkno_list[-6:]
        
        #Setup
        wk_data = []

        my_db = DB_Path()[0]
        my_table = "daily"

        #DB_Read_By_WK
        conn = sqlite3.connect(my_db)
        cur = conn.cursor()

        for wkno in wkno_list:
            cmd = "SELECT MIN(Date) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            querty0 = cur.fetchone()[0]
             
            cmd = "SELECT Open FROM " + my_table + " WHERE Wkno='" + str(wkno) + "' AND Date='" + querty0 +"'"
            cur.execute(cmd)
            op = cur.fetchone()[0]

            cmd = "SELECT MAX(High) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            hp = cur.fetchone()[0]
            
            cmd = "SELECT MIN(Low) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            lp = cur.fetchone()[0]

            cmd = "SELECT MAX(Date) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            querty1 = cur.fetchone()[0]
            
            cmd = "SELECT Close FROM " + my_table + " WHERE Wkno='" + str(wkno) + "' AND Date='" + querty1 +"'"
            cur.execute(cmd)
            cp = cur.fetchone()[0]

            cmd = "SELECT SUM(Turnover) FROM " + my_table + " WHERE Wkno='" + str(wkno)+ "'"
            cur.execute(cmd)
            tos = cur.fetchone()[0]
            
            cmd = "SELECT SUM(Volume) FROM " + my_table + " WHERE Wkno='" + str(wkno)+ "'"
            cur.execute(cmd)
            vs = cur.fetchone()[0]
            
            cmd = "SELECT SUM(Trans) FROM " + my_table + " WHERE Wkno='" + str(wkno)+ "'"
            cur.execute(cmd)
            ts = cur.fetchone()[0]

            wk_data.append([wkno , op , hp , lp , cp , tos , vs , ts])

        conn.close()
        
        DB_Insert("TW" , 2 , wk_data)
        #---------- DONE ----------
        #...........................................................................
        
##############################################################################

class TW_Index_By_Monthly:
    def __init__(self):
        self.DB_data_sum()

    def DB_data_sum(self):
        #...........................................................................
        #Pre Setup
        mon_list = sorted(set([item[3] for item in Read_DB("TW" , 0)]))
        if len(Read_DB("TW" , 3))>0:
            if len(mon_list)>3:
                mon_list = mon_list[-3:]

        #Setup
        mon_data = []

        #DB
        my_db = DB_Path()[0]
        my_table = "daily"
        conn = sqlite3.connect(my_db)
        cur = conn.cursor()

        for mon in mon_list:
            cmd = "SELECT MIN(Date) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            querty0 = cur.fetchone()[0]
            
            cmd = "SELECT Open FROM " + my_table + " WHERE Mon='" + str(mon) + "' AND Date='" + querty0 +"'"
            cur.execute(cmd)
            op = cur.fetchone()[0]

            cmd = "SELECT MAX(High) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            hp = cur.fetchone()[0]
            
            cmd = "SELECT MIN(Low) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            lp = cur.fetchone()[0]

            cmd = "SELECT MAX(Date) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            querty1 = cur.fetchone()[0]
            
            cmd = "SELECT Close FROM " + my_table + " WHERE Mon='" + str(mon) + "' AND Date='" + querty1 +"'"
            cur.execute(cmd)
            cp = cur.fetchone()[0]

            cmd = "SELECT SUM(Turnover) FROM " + my_table + " WHERE Mon='" + str(mon)+ "'"
            cur.execute(cmd)
            tos = cur.fetchone()[0]
            
            cmd = "SELECT SUM(Volume) FROM " + my_table + " WHERE Mon='" + str(mon)+ "'"
            cur.execute(cmd)
            vs = cur.fetchone()[0]
            
            cmd = "SELECT SUM(Trans) FROM " + my_table + " WHERE Mon='" + str(mon)+ "'"
            cur.execute(cmd)
            ts = cur.fetchone()[0]

            mon_data.append([mon , op , hp , lp , cp , tos , vs , ts])

        DB_Insert("TW" , 3 , mon_data)
        #---------- DONE ----------
        #...........................................................................

##############################################################################

class TX_DB_Create:
    def __init__(self):
        self.DB_Create()

    def DB_Create(self):
        my_db = DB_Path()[1]    #["twse.db", "tx.db"]
        #...........................................................................
        # Raw Data DB - Table Lists
        # Table #0 : source
        # Table #1 : working_day
        # Table #2 : contract_list  #契約列表
        # Table #3 : daily
        # Table #4 : weekly
        # Table #5 : monthly
        #...........................................................................
        conn = sqlite3.connect(my_db)
        #...........................................................................
        # <<<<< Table : source >>>>>
        # 日期 週別 年份 月份 開盤指數 最高指數 最低指數  收盤指數
        # 漲跌價 = RF_Price (Rise and Fall in Price)
        # 漲跌幅= RF_Rate 
        # 一般成交量 = Volume 
        # 未沖銷契約量 = OI
        #...........................................................................
        table0 = "source"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table0 + """(
               Date    TEXT    NOT NULL,
               Contract TEXT    NOT NULL,
               Open   TEXT    NOT NULL,
               High     TEXT    NOT NULL,
               Low      TEXT    NOT NULL,
               Close   TEXT    NOT NULL,
               RF_Price TEXT    NOT NULL,
               RF_Rate  TEXT    NOT NULL,
               Volume      INT     NOT NULL,
               OI     INT     NOT NULL,
               UNIQUE(Date , Contract))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table0  + "(Date)"
        conn.execute(cmd)

        #Index - Contract
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Contract ON " + table0  + "(Contract)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        # <<<<< Table : working day >>>>>
        #...........................................................................
        table1 = "working_day"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table1 + """(
               Date    TEXT    NOT NULL,
               Wkno  TEXT    NOT NULL,
               Year     TEXT    NOT NULL,
               Mon     TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table1  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : contract_list >>>>>
        #...........................................................................
        table2 = "contract_list"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table2 + """(
               Contract_list    TEXT    NOT NULL,
               UNIQUE(contract_list))"""
        conn.execute(cmd)
        
        #Index - Contract_list
        cmd = "CREATE INDEX IF NOT EXISTS Idx_contract_list ON " + table2  + "(contract_list)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : daily >>>>>
        # 日期 週別 年份 月份 開盤指數 最高指數 最低指數  收盤指數
        # 一般成交量 = Nor Volume 
        # 盤後成交量 = AH Volume(after-hours trading / 盤後交易)
        # 合計成交量 = Sum Volume
        # 未沖銷契約量 = OI
        #...........................................................................
        table3 = "daily"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table3 + """(
               Date    TEXT    NOT NULL,
               Wkno  TEXT    NOT NULL,
               Year     TEXT    NOT NULL,
               Mon     TEXT    NOT NULL,
               Open   REAL    NOT NULL,
               High     REAL    NOT NULL,
               Low      REAL    NOT NULL,
               Close   REAL    NOT NULL,
               Volume      INT     NOT NULL,
               OI     INT     NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table3  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : weekly >>>>>
        # 週別 開盤指數 最高指數 最低指數  收盤指數
        # 一般成交量 = Volume 
        # 未沖銷契約量 = OI
        #...........................................................................
        table4 = "weekly"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table4 + """(
               Wkno  TEXT    NOT NULL,
               Open   REAL    NOT NULL,
               High     REAL    NOT NULL,
               Low      REAL    NOT NULL,
               Close   REAL    NOT NULL,
               Volume     INT     NOT NULL,
               OI      INT     NOT NULL,
               UNIQUE(Wkno))"""
        conn.execute(cmd)
        
        #Index - Wkno
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Wkno ON " + table4  + "(Wkno)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        # <<<<< Table : monthly >>>>>
        # 年份 開盤指數 最高指數 最低指數  收盤指數
        # 一般成交量 = Volume 
        # 未沖銷契約量 = OI
        #...........................................................................
        table5 = "monthly"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table5 + """(
               Mon     TEXT    NOT NULL,
               Open   REAL    NOT NULL,
               High     REAL    NOT NULL,
               Low      REAL    NOT NULL,
               Close   REAL    NOT NULL,
               Volume     INT     NOT NULL,
               OI      INT     NOT NULL,
               UNIQUE(Mon))"""
        conn.execute(cmd)
        
        #Index - Mon
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Mon ON " + table5  + "(Mon)"
        conn.execute(cmd)
        #...........................................................................
        conn.close()
        #---------- DONE ----------

##############################################################################

class TX_Index_By_Source:
    def __init__(self):
        self.Data_Fetch()
        
    def Data_Fetch(self):
        #...........................................................................
        data_num = len(Read_DB("TX" , 0))
        if data_num==0:
            print("Start Init Set-up For TX")
            self.Init_Update()
        else:
            print("Start Normal Set-up For TX")
        self.Normal_Update()
        self.Date_Handler()
        self.Contract_Handler()
            
        #---------- DONE ----------
        #...........................................................................
            
    def Init_Update(self):
        #...........................................................................
        #加權指數     ※ 本資料自民國88年01月05日開始提供
        #期貨資料 與 加權指數資料 同步
        nyear = int(time.strftime("%Y", time.localtime()))
        ch_dirs = Dir_Chk() #["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
        csv_files = []
        for yy in range(1999 , nyear):
            zip_file = str(yy) + "_fut.zip"
            csv_file = str(yy) + "_fut.csv"
      
            download_path = "http://www.taifex.com.tw/chinese/3/hisdata_fut/" + zip_file

            zip_file = os.path.join(os.path.join(ch_dirs[-1] , str(yy))  , zip_file)

            time.sleep(3)
            urllib.request.urlretrieve(download_path , zip_file)

            #Unzip
            with zipfile.ZipFile(zip_file) as f:
                f.extract(csv_file, os.path.join(ch_dirs[-1] , str(yy)))
                csv_files.append(os.path.join(os.path.join(ch_dirs[-1] , str(yy))  , csv_file))
                
        #Include [Fun]DB_Insert
        self.csv_read(csv_files)

        #Remove csv_file
        for csv_file in csv_files:
            os.remove(csv_file)
        #---------- DONE ----------
        #...........................................................................

    
            
    def csv_read(self , csv_files):
        #...........................................................................
        source_data = []
        format_dict = {1 : '交易日期' , 2 : '契約' , 3 : '到期月份(週別)' ,  4 : '開盤價' , 5 : '最高價' , 6 : '最低價' ,
                       7 : '收盤價' , 8 : '漲跌價' , 9 : '漲跌%' ,  10 : '成交量' , 11 : '結算價' , 12 : '未沖銷契約數' ,
                       13 : '最後最佳買價' , 14 : '最後最佳賣價' , 15 : '歷史最高價' , 16 : '歷史最低價' , 17 : '是否因訊息面暫停交易' , 18 : '交易時段'}
        #------------------------------------------
        for csv_file in csv_files:
            fp = open(csv_file , "r")
            
            csv_fp = csv.reader(fp)
            
            #print("CSV File : " , csv_file ," - Checking")
            format_no = 0
            
            for count , content in enumerate(csv_fp):
                temp = []                
                if count==0:
                    row_count = len(content)
                    for i in range(0 , row_count):
                        if content[i]==format_dict[i+1]:
                            #print("欄位" + str(i+1) + " : " + format_dict[i+1] + " - 格式正確")
                            format_no += 1
                        else:
                            #print("欄位" + str(i+1) + " : " + format_dict[i+1] + " - 格式不正確")
                            word = "欄位" + str(i+1) + " : " + format_dict[i+1] + " - 格式不正確"                            
                else:
                    #1998/7/21
                    t = time.strptime(content[0] , "%Y/%m/%d")
                    row_date = time.strftime("%Y-%m-%d" , t)
                    
                    if format_no<18:
                        if content[1]=="TX" and len(content[2])==6:
                            temp.append(row_date)
                            for i in range(2,10):temp.append(content[i])
                            temp.append(content[11])
                            
                    elif format_no==18:
                        if content[1]=="TX" and content[-1]=="一般" and len(content[2])==6:
                            temp.append(row_date)
                            for i in range(2,10):
                                temp.append(content[i])
                            temp.append(content[11])
                    else:
                        print("Error - Csv Format Fail")
                        print("Title Format Row Error...Please Check!!!!")
                        print("Stop Procedure!")
                        sys.exit(0)

                if len(temp)>0:
                    source_data.append(temp)
            #------------------------------------------
        DB_Insert("TX" ,  0 , source_data)
        #---------- DONE ----------
        #...........................................................................

    def Normal_Update(self):
        #...........................................................................
        insert_data = []
        sql_querty1 = Read_DB("TW" , 0)
        sql_querty2 = Read_DB("TX" , 0)

        #Get Difference-Dates Between TWSE & TX
        tw_dates = sorted(set([item[0] for item in sql_querty1]))
        tx_dates = sorted(set([item[0] for item in sql_querty2]))
        updates = sorted(list(set(tw_dates).difference(set(tx_dates))))

        #Start Url Fetch
        for dd in updates:
            #Format 2017-01-03
            self.tx_fetch(dd , insert_data)
        DB_Insert("TX" , 0 , insert_data)
        #---------- DONE ----------
        #...........................................................................
            
    def tx_fetch(self , dates , insert_data):
        #...........................................................................      
        ceDates = dates.split("-")
        
        #my_url = "http://www.taifex.com.tw/chinese/3/3_1_1.asp"
        my_url = "http://www.taifex.com.tw/cht/3/futDailyMarketReport"
        
        my_data = {
            "syear" : ceDates[0] ,
            "smonth" : ceDates[1] ,
            "sday" : ceDates[2] ,
            "market_code" : "0" ,
            "datestart" : dates.replace("-" , "/"),
            "commodity_idt" : "TX",
            }

        # <<< RE Function >>>
        r = Url_Fetch(my_url , my_data)
        r.encoding = "utf-8"
        content = r.text
        #print(content)
        
        #<Date : 2017/10/12>
        pattern0 = r"""<h3 align="left">日期：(\d+/\d+/\d+)</h3>"""
        reg0 = re.compile(pattern0)
        match0 = reg0.findall(content)
        if len(match0)==0:
            print("[Err]TX URL1 DATE FETCH ERROR!!!")
            input("PLEASE ENTET TO LEAVE")
            sys.exit(0)

        #<Table Data >
        #"到期月份(週別)" , "開盤價" , "最高價" ,"最低價" ,"最後成交價" ,"漲跌價" , "漲跌%" , "*盤後交易時段成交量" , "*一般交易時段成交量" , "*合計成交量" , "結算價" , "*未沖銷契約量" , "最後最佳買價" , "最後最佳賣價" , "歷史最高價"
        #[('201710', '10657', '10705', '10640', '10697', '▲65', '▲ 0.61%', '5864', '103406', '109270', '10697', '89931', '10697', '10698', '10705', '10186'),
        #('201711', '10636', '10688', '10625', '10682', '▲67', '▲ 0.63%', '389', '10793', '11182', '10682', '15893', '10682', '10683', '10688', '10241'),
        #('201712', '10627', '10672', '10609', '10666', '▲66', '▲ 0.62%', '60', '613', '673', '10666', '10024', '10665', '10666', '10672', '8856'),
        #('201803', '10567', '10619', '10556', '10612', '▲67', '▲ 0.64%', '94', '751', '845', '10612', '2567', '10611', '10613', '10619', '9146'),
        #('201806', '10502', '10553', '10492', '10548', '▲67', '▲ 0.64%', '9', '364', '373', '10548', '1093', '10547', '10550', '10553', '9456')]
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk"><div align="center">(\d\d\d\d\d\d)\s*</div></td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="middle" class="12green">\s*\S*<FONT color=\w+>(\S*\s*)</FONT>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<FONT color=\w+>(\S*\s*\S*)</FONT>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*\S*(\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*"""
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk"><div align="center">(\d\d\d\d\d\d)\s*</div></td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="middle" class="12green">\s*\S*<FONT color=\w+>(\S*\s*)</FONT>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<FONT color=\w+>(\S*\s*\S*)</FONT>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*\S*(\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*"""
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk"><div align="center">(\d\d\d\d\d\d)\s*</div></td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="middle" class="12green">\s*\S*<font color="\w+">(\S*\s*)</font>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<font color="\w+">(\S*\s*\S*)</font>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*\S*(\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\S*"""
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""

        #Ver 1.1.16B
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""
                
        #Ver 1.1.16C
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*([\S*\d+|-])\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""

        #Ver 1.1.16D
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*(\S*-)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""
        #pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color=.*\w+.*>(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color=.*\w+.*>(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""
   
        RE_Counter = 0
        #print(content)
        fin_match = []
        t11 = time.time()

        #Ver 1.1.21
        if True:
            print("[TX Fetch]Condition1 - Execute")
            
            #Ver 1.1.21 - Verify OK
            #pattern1 = r"""\s*<td class="12bk">\s*\s*<div align="center">TX\s*</div>\s*\s*</td>\s*\s*<td class="12bk">\s*\s*<div align="center">(\d\d\d\d\d\d)\s*</div></td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="middle" class="12green">\s*\s*<\w\w\w\w color=\w+>(\S*\d+)</\w\w\w\w>\s*\s*</td>\s*\s*<td align="middle" class="12green">\s*\s*<\w\w\w\w color=\w+>(\S*\d+%)</\w\w\w\w>\s*\s*</td>\s*"""
            pattern1 = r"""\s*<td class="12bk">\s*\s*<div align="center">TX\s*</div>\s*\s*</td>\s*\s*<td class="12bk">\s*\s*<div align="center">(\d\d\d\d\d\d)\s*</div></td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\S*\d+)\s*</td>\s*\s*<td align="middle" class="12green">\s*\s*<\w\w\w\w color=\S*\w+\S*>(\S*\d+)</\w\w\w\w>\s*\s*</td>\s*\s*<td align="middle" class="12green">\s*\s*<\w\w\w\w color=\S*\w+\S*>(\S*\d+%)</\w\w\w\w>\s*\s*</td>\s*"""
            pattern1 += r"""\s*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\s*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\s*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*"""
            pattern1 += r"""\s*<td align="right" class="12bk">\s*\s*(\S*)</td>\s*\s*<td align="right" class="12bk">\s*\s*(\d+)\s*</td>\s*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\s*<td align="right" class="12bk">\s*(\d+)\s*</td>\s*\s*</tr>"""

            reg1 = re.compile(pattern1) 
            match1 = reg1.findall(content)
            for item in match1:
                #print(item)
                fin_match.append(item)

        if len(fin_match)!=6:
            print("[TX Fetch]Condition1 - FAIL\n")

            print("[TX Fetch]Condition2 - Execute")
            fin_match = []
            for i in range(2):
                if i==0:pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*(\S*-)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""
                else:pattern1 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color=.*\w+.*>(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color=.*\w+.*>(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""
                reg1 = re.compile(pattern1)       
                match1 = reg1.findall(content)
                for item in match1:
                    #print(item)
                    fin_match.append(item)

        if len(fin_match)!=6:
            print("[TX Fetch]Condition2 - FAIL\n")

            print("[TX Fetch]Condition3 - Execute")
            
            #Ver 1.1.16B
            pattern2 = r"""<td class="12bk">\s*\S*<div align="center">TX\s*\S*</div>\s*\S*</td>\s*\S*<td class="12bk">\s*\S*<div align="center">(\d\d\d\d\d\d)\s*</div>\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<td align="middle" class="12green">\s*\S*<\w+ color="\w+">(\S*\s*\S*)</\w+>\s*\S*</td>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<TD align=right class="12bk">\s*(\d+)\s*</TD>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*<td align="right" class="12bk">\s*(\S*\d+)\s*\S*</td>\s*\S*"""
            reg2 = re.compile(pattern2)
            match1 = reg2.findall(content)

            if len(match1)==0:
                #print(content)
                print("[Err]TX URL DATA FETCH ERROR!!!")
                print("RE-PATTERN1 and RE-PATTERN2 and RE-PATTERN3 - EXECUTE FAILURE !")
                input("PLEASE ENTET TO LEAVE")
                sys.exit(0)

        insert_data = self.fetch_data_handle(dates , match0 , fin_match , insert_data)
        return(insert_data)
        #---------- DONE ----------
        #...........................................................................

    def fetch_data_handle(self , dates , data1 , data2 , insert_data):
        #...........................................................................
        #dates : Format 2017-01-03 
        #data1 : Format 2017/10/12 (From Fetch)
        fetch_date = data1[0].replace("/" , "-")
        if dates==fetch_date:
            for row in sorted(data2):
                temp = [dates , row[0] , row[1] , row[2] , row[3] , row[4] , row[5][1:] , row[6][1:] , row[8] , row[11]]
                insert_data.append(temp)
                
            self.fetch_data_save(dates , data2)
        return(insert_data)
        #---------- DONE ----------
        #...........................................................................
                

    def fetch_data_save(self , dates , fetch_data):
        #...........................................................................
        ceDates = dates.split("-")

        title = ["契約" , "到期月份(週別)" , "開盤價" , "最高價" ,"最低價" ,"最後成交價" ,"漲跌價" , "漲跌%" , "*盤後交易時段成交量" , "*一般交易時段成交量" , "*合計成交量" , "結算價" , "*未沖銷契約量" , "最後最佳買價" , "最後最佳賣價" , "歷史最高價"]

        file_name = "".join(ceDates) + "_fut.csv"
        
        #ch_dirs = ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
        year = ceDates[0]
        mon = ceDates[1]
        
        ch_dir1 = os.path.join(Dir_Chk()[-1]  , str(year))
        ch_dir2 = os.path.join(ch_dir1  , str(year)+ str(mon).zfill(2))
            
        csv_file = os.path.join(ch_dir2 , file_name)

        #原始資料整理
        csv_content = [item for item in fetch_data]
        csv_content.insert(0 , title)
        
        with open(csv_file , 'w') as fp:
            f = csv.writer(fp)
            f.writerows(csv_content)   
        fp.close()
        #---------- DONE ----------
        #...........................................................................
        
    def Date_Handler(self):
        #...........................................................................
        sql_querty = Read_DB("TX" , 0)
        dates_list = [item[0] for item in sql_querty]
        content = []

        for d in dates_list:
            #ceDates = d.split("/")
            t = time.strptime(d , "%Y-%m-%d")
            wkno = time.strftime("%Y%U" , t)
            
            if int(wkno)==0:
                t = time.strptime(str(int(y)-1)+"/12/31" , "%Y/%m/%d")
                wkno = time.strftime("%YWK%U" , t)
            else:
                wkno = time.strftime("%YWK%U" , t)
            #d = ceDates[0] + "-" + ceDates[1].zfill(2) + "-" + ceDates[2].zfill(2)
            temp = [d , wkno , time.strftime("%Y" , t) , time.strftime("%Y%m" , t)]
            content.append(temp)
            
        DB_Insert("TX" , 1 , content)
        #---------- DONE ----------
        #...........................................................................
        
    def Contract_Handler(self):
        #...........................................................................
        content = []
        
        sql_querty = Read_DB("TX" , 0)
        
        contract_list = sorted(set([item[1] for item in sql_querty]))
        
        if len(Read_DB("TX" , 2))>0:
            contract_list = contract_list[-12:]
        
        for c in contract_list:
            content.insert(-1 , [c])
            
        DB_Insert("TX" , 2 , content)
        #---------- DONE ----------
        #...........................................................................

########################################################################################################

class TX_Index_By_Daily:
    def __init__(self):
        self.DB_data_sum()

    def DB_data_sum(self):
        #...........................................................................
        #Set-up
        daily_data = []

        # Table #1 : working_day
        # Table #3 : daily
        sql_querty1 = Read_DB("TX" , 1)
        sql_querty2 = Read_DB("TX" , 3)
        if len(sql_querty2)>0: sql_querty1[-90:]

        #DB_Read_By_Source
        #-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-
        my_db = DB_Path()[1]
        my_table = "source"
        conn = sqlite3.connect(my_db)
        cur = conn.cursor()
        
        for date , wkno , year , mon in sql_querty1:
            cmd = "SELECT MIN(Contract) FROM " + my_table + " WHERE Date='" + str(date ) + "'"
            cur.execute(cmd)
            contract_querty = cur.fetchone()[0]

            cmd = "SELECT * FROM " + my_table + " WHERE Date='" + str(date ) + "' AND Contract='" + contract_querty + "'"
            cur.execute(cmd)
            source_querty = cur.fetchall()[0]

            daily_data.append([date , wkno , year , mon , source_querty[2] , source_querty[3] , source_querty[4] , source_querty[5] , source_querty[8] , source_querty[9]])        
        conn.close()
         #-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-
        DB_Insert("TX" , 3 , daily_data)
        #---------- DONE ----------
        #...........................................................................

########################################################################################################

class TX_Index_By_Weekly:
    def __init__(self):
        self.DB_data_sum()

    def DB_data_sum(self):
        #...........................................................................
        # Table #3 : daily
        # Table #4 : weekly
        sql_querty = Read_DB("TX" , 4)
        wkno_list = sorted(set(item[1] for item in (Read_DB("TX" , 3))))
        
        if len(sql_querty)>0:
            wkno_list = wkno_list[-12:]
        
        #Set-up
        wk_data = []

        #DB
        my_db = DB_Path()[1]
        my_table = "daily"
        conn = sqlite3.connect(my_db)
        cur = conn.cursor()

        for wkno in wkno_list:
            cmd = "SELECT MIN(Date) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            querty0 = cur.fetchone()[0]
            
            cmd = "SELECT Open FROM " + my_table + " WHERE Wkno='" + str(wkno) + "' AND Date='" + querty0 +"'"
            cur.execute(cmd)
            op = cur.fetchone()[0]

            cmd = "SELECT MAX(High) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            hp = cur.fetchone()[0]
            
            cmd = "SELECT MIN(Low) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            lp = cur.fetchone()[0]

            cmd = "SELECT MAX(Date) FROM " + my_table + " WHERE Wkno='" + str(wkno) + "'"
            cur.execute(cmd)
            querty1 = cur.fetchone()[0]
            
            cmd = "SELECT Close FROM " + my_table + " WHERE Wkno='" + str(wkno) + "' AND Date='" + querty1 +"'"
            cur.execute(cmd)
            cp = cur.fetchone()[0]
            
            cmd = "SELECT SUM(Volume) FROM " + my_table + " WHERE Wkno='" + str(wkno)+ "'"
            cur.execute(cmd)
            vol = cur.fetchone()[0]
            
            cmd = "SELECT SUM(OI) FROM " + my_table + " WHERE Wkno='" + str(wkno)+ "'"
            cur.execute(cmd)
            oi = cur.fetchone()[0]

            wk_data.append([wkno , op , hp , lp , cp , vol , oi])
        conn.close()
         #-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-
        DB_Insert("TX" , 4 , wk_data)
        #---------- DONE ----------
        #...........................................................................

########################################################################################################

class TX_Index_By_Monthly:
    def __init__(self):
        self.DB_data_sum()

    def DB_data_sum(self):
        #...........................................................................
        # Table #3 : daily
        # Table #5 : monthly
        sql_querty = Read_DB("TX" , 5)
        mon_list = sorted(set(item[3] for item in (Read_DB("TX" , 3))))
        
        if len(sql_querty)>0:
            mon_list = mon_list[-3:]
        
        #Set-up
        mon_data = []

        #DB
        my_db = DB_Path()[1]
        my_table = "daily"
        conn = sqlite3.connect(my_db)
        cur = conn.cursor()

        for mon in mon_list:
            cmd = "SELECT MIN(Date) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            querty0 = cur.fetchone()[0]
            
            cmd = "SELECT Open FROM " + my_table + " WHERE Mon='" + str(mon) + "' AND Date='" + querty0 +"'"
            cur.execute(cmd)
            op = cur.fetchone()[0]

            cmd = "SELECT MAX(High) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            hp = cur.fetchone()[0]
            
            cmd = "SELECT MIN(Low) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            lp = cur.fetchone()[0]

            cmd = "SELECT MAX(Date) FROM " + my_table + " WHERE Mon='" + str(mon) + "'"
            cur.execute(cmd)
            querty1 = cur.fetchone()[0]
            
            cmd = "SELECT Close FROM " + my_table + " WHERE Mon='" + str(mon) + "' AND Date='" + querty1 +"'"
            cur.execute(cmd)
            cp = cur.fetchone()[0]
            
            cmd = "SELECT SUM(Volume) FROM " + my_table + " WHERE Mon='" + str(mon)+ "'"
            cur.execute(cmd)
            vol = cur.fetchone()[0]
            
            cmd = "SELECT SUM(OI) FROM " + my_table + " WHERE Mon='" + str(mon)+ "'"
            cur.execute(cmd)
            oi = cur.fetchone()[0]

            mon_data.append([mon , op , hp , lp , cp , vol , oi])
        conn.close()
         #-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-
        DB_Insert("TX" , 5 , mon_data)
        #---------- DONE ----------
        #...........................................................................
        
########################################################################################################


class TXO_DB_Create:
    def __init__(self):
        self.DB_Create()

    def DB_Create(self):
        my_db = DB_Path()[2]    #["twse.db", "tx.db" ,  "txo.db"]
        #...........................................................................
        # Raw Data DB - Table Lists
        # Table #0 : 
        #...........................................................................
        conn = sqlite3.connect(my_db)
        #...........................................................................
        # <<<<< Table : Settlement_List >>>>>
        #...........................................................................
        table0 = "settlement_list"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table0 + """(
               Date    TEXT    NOT NULL,
               Contract_List  TEXT    NOT NULL,
               Settlement_Price     TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table0  + "(Date)"
        conn.execute(cmd)
        
        #Index - Contract
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Contract_List ON " + table0  + "(Contract_List)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : working day >>>>>
        #...........................................................................
        table1 = "working_day"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table1 + """(
               Date    TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table1  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        conn.close()
        #---------- DONE ----------
        #...........................................................................

########################################################################################################

class TXO:
    def __init__(self):
        self.Data_Fetch()

    def Data_Fetch(self):
        #...........................................................................
        data_num = len(Read_DB("TXO" , 0))
        if data_num==0:
            print("Start Init Set-up For TXO")
            self.Init_Update_For_Settlement()
            self.Init_Update_For_Option()
            self.Normal_Update_For_Option()
            
        else:
            print("Start Normal Set-up For TXO")
            self.Normal_Update_For_Settlement()
            self.Normal_Update_For_Option()
        #---------- DONE ----------
        #...........................................................................

    def Init_Update_For_Settlement(self):
        #...........................................................................
        #my_url = "http://www.taifex.com.tw/chinese/5/FutIndxFSP.asp"
        my_url = "http://www.taifex.com.tw/cht/5/futIndxFSP"

        my_data = {
                   "start_year" : "2001" ,
                   "start_month" : "01" ,
                   "end_year" :  time.strftime("%Y", time.localtime()) ,
                   "end_month" :  time.strftime("%m", time.localtime()) ,
                   "_all" : "on" ,
                   "button" : "送出查詢" ,
                   "commodityIds" :"1" ,
                   }

        # <<< RE Function >>>
        r = Url_Fetch(my_url , my_data)
        r.encoding = "utf-8"
        content = r.text
        
        pattern = r"""\s*<TD width="14%" align=middle>(\d+/\d+/\d+)</TD>\s*\s*<TD width="10%" align=middle>(\S+)\s*</TD>\s*<TD width="10%" align=right>(\d*)</TD>\s*"""
        
        reg = re.compile(pattern)
        
        match = reg.findall(content)
        
        if len(match)==0:
            print("[Err]TXO-Settlement FETCH ERROR!!!")
            input("PLEASE ENTET TO LEAVE")
            sys.exit(0)
            
        insert_data = []
        
        for i , j , k in match:
            my_date = i.split("/")
            insert_data.append([my_date[0] + "-"+my_date[1].zfill(2) + "-" + my_date[2].zfill(2) , j.strip() ,k])
        
        #insert_data = [ item for item in sorted(set(match))]

        DB_Insert("TXO" , 0 , insert_data)
        #---------- DONE ----------
        #...........................................................................

    def Normal_Update_For_Settlement(self):
        #...........................................................................
        nyear = time.strftime("%Y", time.localtime())
        nmon = time.strftime("%m", time.localtime())
        if int(nmon)==1:
            syear = str(int(nyear)-1)
            smon = "12"
            eyear = nyear
            emon = nyear
 
        else:
            syear = nyear
            smon = str(int(nmon)-1).zfill(2)
            eyear = nyear
            emon = nyear
        
        my_url = "http://www.taifex.com.tw/cht/5/futIndxFSP"

        my_data = {
                   "start_year" : syear  ,
                   "start_month" : smon  ,
                   "end_year" : eyear  ,
                   "end_month" : emon  ,
                   "_all" : "on" ,
                   "button" : "送出查詢" ,
                   "commodityIds" :"1" ,
                   }

        # <<< RE Function >>>
        r = Url_Fetch(my_url , my_data)
        r.encoding = "utf-8"
        content = r.text

        #pattern = """<TR bgcolor="#FFFFFF" class="12bk">\s*\S*<TD width="14%" align=middle>(\d+/\d+/\d+)</TD>\s*\S*<TD width="10%" align=middle>(\d+\D*\d*)\s*</TD>\s*\S*<TD width="10%" align=right>(\d*)\s*</TD>\s*\S*"""
        pattern = r"""\s*<TD width="14%" align=middle>(\d+/\d+/\d+)</TD>\s*\s*<TD width="10%" align=middle>(\S+)\s*</TD>\s*<TD width="10%" align=right>(\d*)</TD>\s*"""
        
        reg = re.compile(pattern)
        
        match = reg.findall(content)

        if len(match)==0:
            print("[Err]TXO-Settlement FETCH ERROR!!!")
            input("PLEASE ENTET TO LEAVE")
            sys.exit(0)

        insert_data = []
        
        for i , j , k in match:
            my_date = i.split("/")
            insert_data.append([my_date[0] + "-"+my_date[1].zfill(2) + "-" + my_date[2].zfill(2) , j.strip() ,k])

        DB_Insert("TXO" , 0 , insert_data)
        #---------- DONE ----------
        #...........................................................................

    def Init_Update_For_Option(self):
        #...........................................................................
        nyear = int(time.strftime("%Y", time.localtime()))
        
        ch_dirs = Dir_Chk() #["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
        
        txo_csv_files = []

        for yy in range(2001 , nyear):
            zip_file = str(yy) + "_opt.zip"
                        
            #http://www.taifex.com.tw/chinese/3/hisdata_opt/2001_opt.zip
            #http://www.taifex.com.tw/cht/3/optDailyMarketView
            download_path = "http://www.taifex.com.tw/chinese/3/hisdata_opt/" + zip_file

            zip_file = os.path.join(os.path.join(ch_dirs[-1] , str(yy))  , zip_file)

            time.sleep(3)
            
            urllib.request.urlretrieve(download_path , zip_file)

            with zipfile.ZipFile(open(zip_file, 'rb')) as f:
                for file in f.namelist():
                    f.extract(file, os.path.join(ch_dirs[-1] , str(yy)))
                    txo_csv_files.append(os.path.join(os.path.join(ch_dirs[-1] , str(yy))  , file))
            f.close()

        #<<< DB Insert >>>
        #return(txo_csv_files)
        source_data = self.Csv_Read(txo_csv_files)

        #DB Insert Function
        self.Source_Data_Handle(source_data)
        #---------- DONE ----------
        #...........................................................................

    def Csv_Read(self , csv_files):
        #---------------------------------------------------------------------------------------------------------------------------------
        #..................................................................................................................................................................................................................................
        source_data = []
        format_dict = { 1 : '交易日期' , 2 : '契約' , 3 : '到期月份(週別)' , 4 : ' 履約價' ,
                        5 : '買賣權' , 6 : '開盤價' , 7 : '最高價' , 8 : '最低價' ,
                        9 : '收盤價' , 10 : '成交量' , 11 : '結算價' , 12 : '未沖銷契約數' ,
                        13 : '最後最佳買價' , 14 : '最後最佳賣價' , 15 : '歷史最高價' , 16 : '歷史最低價' ,
                        17 : '是否因訊息面暫停交易' , 18 : '交易時段'}
        #..................................................................................................................................................................................................................................
        for csv_file in csv_files:
            
            fp = open(csv_file , "r")
            csv_fp = csv.reader(fp)
            format_no = 0
                
            for count , content in enumerate(csv_fp):
                temp = []                
                if count==0:
                    row_count = len(content)
                    for i in range(0 , row_count):
                        if content[i]==format_dict[i+1]:
                            format_no += 1
                        else:
                            word = "欄位" + str(i+1) + " : " + format_dict[i+1] + " - 格式不正確"
                else:
                    t = time.strptime(content[0] , "%Y/%m/%d")
                    row_date = time.strftime("%Y-%m-%d" , t)
                        
                    if format_no<18:
                        if content[1]=="TXO" and len(content[2])<=8:
                            temp.append(row_date)
                            for i in range(2,10):temp.append(content[i])
                            temp.append(content[11])
                                
                    elif format_no==18:
                        if content[1]=="TXO" and content[-1]=="一般" and len(content[2])<=8:
                            temp.append(row_date)
                            for i in range(2,10):
                                temp.append(content[i])
                            temp.append(content[11])
                    else:
                        print("Error - Csv Format Fail")
                        print("Title Format Row Error...Please Check!!!!")
                        print("Stop Procedure!")
                        sys.exit(0)

                    if len(temp)>0:
                        source_data.append(temp)
            fp.close()
            os.remove(csv_file)    
        return(source_data)
        #---------- DONE ----------
        #---------------------------------------------------------------------------------------------------------------------------------
    
    def Source_Data_Handle(self , source_data):
        #---------------------------------------------------------------------------------------------------------------------------------
        expiry_dates = sorted(set([s[1] for s in source_data]))
        #---------------------------------------------------------------------------------------------------------------------------------
        for expiry_date in expiry_dates:
            insert_data = []
            for content in source_data:
                #print(content)
                if content[1]==expiry_date:
                    insert_data.append([content[0] , content[2], content[3], content[4], content[5], content[6], content[7], content[8], content[9]])
                        
            #print(len(insert_data))
            self.TXO_DB_Handle(expiry_date , insert_data)
        #---------- DONE ----------
        #---------------------------------------------------------------------------------------------------------------------------------


    def TXO_DB_Handle(self , expiry_date , insert_data):
        #---------------------------------------------------------------------------------------------------------------------------------
        #Format : expiry_date = 200201
        db_dir = Dir_Chk()[-2] # ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
        my_db = os.path.join(db_dir , "TXO" + expiry_date[:4] +".db.")
        my_table = "OP_"+expiry_date
        
        conn = sqlite3.connect(my_db)
        #Strike price : 履約價
        cmd = "CREATE TABLE IF NOT EXISTS " + my_table + """(
                   Date    TEXT    NOT NULL,
                   Strike_Price TEXT    NOT NULL,
                   BS_Type TEXT    NOT NULL,
                   Open   TEXT    NOT NULL,
                   High     TEXT    NOT NULL,
                   Low      TEXT    NOT NULL,
                   Close   TEXT    NOT NULL,
                   Volume      INT     NOT NULL,
                   OI     INT     NOT NULL,
                   UNIQUE(Date , Strike_Price , BS_Type))"""
        conn.execute(cmd)

        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + my_table  + "(Date)"
        conn.execute(cmd)

        #Index - Strike_Price
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Strike_Price ON " + my_table  + "(Strike_Price)"
        conn.execute(cmd)

        #Index - BS_Type
        cmd = "CREATE INDEX IF NOT EXISTS Idx_BS_Type ON " + my_table  + "(BS_Type)"
        conn.execute(cmd)
        #..................................................................................................................................................................................................................................

        #Insert Data
        cmd = "INSERT OR REPLACE INTO " + my_table + " VALUES(?,?,?,?,?,?,?,?,?)"
        
        conn.executemany(cmd , insert_data)
        
        conn.commit()
        
        conn.close()
        #---------------------------------------------------------------------------------------------------------------------------------

    def  Normal_Update_For_Option(self):
        #---------------------------------------------------------------------------------------------------------------------------------
        sql_querty1 = Read_DB("TX" , 1)
        sql_querty2 = Read_DB("TXO" , 1)
        
        yy = time.strftime("%Y" , time.localtime())
        
        source_data = []

        dates_list = []
        
        if len(sql_querty2)==0:
            updates = sorted(set([item[0] for item in sql_querty1 if str(item[2])==yy]))
            dates_list = [ [item[0]] for item in sorted(set(sql_querty1))]
            
        else:
            tx_dates = sorted(set([item[0] for item in sql_querty1]))
            txo_dates = sorted(set([item[0] for item in sql_querty2]))
            updates = sorted(list(set(tx_dates).difference(set(txo_dates))))
            dates_list = [ [item] for item in sorted(set(updates))]
    
        for dd in sorted(set(updates)):
            #Format : '2017-01-03'
            myDate = dd.split("-")
             
            #my_url = "http://www.taifex.com.tw/chinese/3/3_2_2.asp"
            my_url = "http://www.taifex.com.tw/cht/3/optDailyMarketReport"
            
            my_data = {
                "commodity_id" : "TXO" ,
                "commodity_idt" : "TXO" ,
                "marketCode" : "0" ,
                "MarketCode" : "0" ,
                "queryDate" : dd.replace("-" , "/") ,
                "queryType" : "2" ,
                }
            
            # <<< RE Function >>>
            r = Url_Fetch(my_url , my_data)
            
            (source_data , raw_data) = self.Fetch_Data_Handle(dd , r , source_data)
            
            self.Fetch_Data_Save(dd , raw_data)
            
        self.Source_Data_Handle(source_data)
        
        DB_Insert("TXO" , 1 , dates_list)
        #---------- DONE ----------
        #---------------------------------------------------------------------------------------------------------------------------------

    def Fetch_Data_Save(self , dates , fetch_data):
        #---------------------------------------------------------------------------------------------------------------------------------
        title = ["日期" , "到期月份(週別)" , "履約價" , "買賣權" , "開盤價" , "最高價" ,"最低價" ,"最後成交價" , "結算價" , "*盤後交易時段成交量" , "一般交易時段成交量" , "*合計成交量" , "未沖銷契約量" ,  "最後最佳買價" , "最後最佳賣價" , "歷史最高價" , "歷史最低價"]

        ceDates = dates.split("-")

        file_name = "".join(ceDates) + "_txo.csv"
        
        #ch_dirs = ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
        year = ceDates[0]
        mon = ceDates[1]
        
        ch_dir1 = os.path.join(Dir_Chk()[-1]  , str(year))
        ch_dir2 = os.path.join(ch_dir1  , str(year)+ str(mon).zfill(2))
            
        csv_file = os.path.join(ch_dir2 , file_name)

        #原始資料整理
        csv_content = [item for item in fetch_data]
        csv_content.insert(0 , title)
        
        with open(csv_file , 'w') as fp:
            f = csv.writer(fp)
            f.writerows(csv_content)   
        fp.close()
        #---------- DONE ----------
        #...........................................................................

    def Fetch_Data_Handle(self , myDate , r , source_data):
        #---------------------------------------------------------------------------------------------------------------------------------
        r.encoding = "utf-8"
        
        content = r.text
        
        pattern0 = """\s*<td class="12bk">\s*\s*<div align="center">TXO\s*</div>\s*\s*</td>\s*\s*<td class="12bk"><div align="center">(\S+)\s*</div></td>\s*"""
        reg0 = re.compile(pattern0)
        match0 = reg0.findall(content)

        pattern = r"""\s*<div align="center">(\D\D\D)\s*</div>\s*\s*</td>\s*\s*<td class="12bk"><div align="center">(\d\d\d\d\d\d\D*\d*)</div></td>\s*\s*<td align="right" class="12bk">(\d+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">\s*\s*(\S+)</td>\s*\s*<td align="middle" class="12green">\s*\s*\S*\s*\S*\s*\s*\s*</td>\s*\s*<td align="middle" class="12green">\s*\s*\S*\s*\S*\s*\s*\s*</td>\s*\s*<TD align=right class="12bk">(\d+)</TD>\s*\s*<TD align=right class="12bk">(\d+)</TD>\s*\s*<TD align=right class="12bk">(\d+)</TD>\s*\s*<TD align=right class="12bk">(\d+)</TD>\s*\s*</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*<td align="right" class="12bk">(\S+)</td>\s*\s*</tr>\s*"""
        reg = re.compile(pattern)
        match1 = reg.findall(content)
        
        if len(match0)==0 or len(match1)==0:
            print("[Err]TXO DATA FETCH ERROR!!!")
            input("PLEASE ENTET TO LEAVE")
            sys.exit(0)
            
        #print(match)
        #print(len(match0) , len(match))
        #input("XXXx")

        if not len(match0)==len(match1):
            print("TXO Fetch Counter Is Not Macth : " , myDate)

            print("Please Confirm RE Rule!")
            print("Stop Procedure!")
            input("Please Enter Any To Leave!!!")  
            sys.exit(0)

        for item in match1:
            if item[3].strip()=="Call":
                bs_type="買權"
            elif item[3].strip()=="Put":
                bs_type="賣權"
            #temp = [myDate , item[1] , item[2] , item[3] , item[4] , item[5] , item[6] , item[7] , item[10] , item[12]]
            source_data.append([myDate , item[1] , item[2] , bs_type , item[4] , item[5] , item[6] , item[7] , item[10] , item[12]])
            
        return(source_data , match1)
        #---------- DONE ----------
        #---------------------------------------------------------------------------------------------------------------------------------

########################################################################################################
class TXF_DB_Create:
    def __init__(self):
        self.DB_Create()

    def DB_Create(self):
        my_db = DB_Path()[3]    #["twse.db", "tx.db" , "txo.db" , "txf.db"]
        #...........................................................................
        # Raw Data DB - Table Lists
        # Table #0 : working_day
        # Table #1 : ii1                     * 自營商
        # Table #2 : ii2                     * 投信
        # Table #3 : ii3                     * 外資
        #...........................................................................
        conn = sqlite3.connect(my_db)
        #...........................................................................
        # <<<<< Table : working day >>>>>
        #...........................................................................
        table0 = "working_day"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table0 + """(
               Date    TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table0  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii1 >>>>>
        # Date
        # Long OI = 多方OI
        # Long OI Money = 多方OI 金額
        # Short OI = 空方OI
        # Short OI Money = 空方OI 金額
        # Diff OI = 多空方OI差
        # Diff OI Money = 多空方OI 金額差
        #...........................................................................
        table1 = "ii1"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table1 + """(
               Date    TEXT    NOT NULL,
               Long_OI  TEXT    NOT NULL,
               Long_OI_Money  TEXT    NOT NULL,
               Short_OI  TEXT    NOT NULL,
               Short_OI_Money  TEXT    NOT NULL,
               Diff_OI  TEXT    NOT NULL,
               Diff_OI_Money  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table1  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii2 >>>>>
        #...........................................................................
        table2 = "ii2"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table2 + """(
               Date    TEXT    NOT NULL,
               Long_OI  TEXT    NOT NULL,
               Long_OI_Money  TEXT    NOT NULL,
               Short_OI  TEXT    NOT NULL,
               Short_OI_Money  TEXT    NOT NULL,
               Diff_OI  TEXT    NOT NULL,
               Diff_OI_Money  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table2  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii3 >>>>>
        #...........................................................................
        table3 = "ii3"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table3 + """(
               Date    TEXT    NOT NULL,
               Long_OI  TEXT    NOT NULL,
               Long_OI_Money  TEXT    NOT NULL,
               Short_OI  TEXT    NOT NULL,
               Short_OI_Money  TEXT    NOT NULL,
               Diff_OI  TEXT    NOT NULL,
               Diff_OI_Money  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table3  + "(Date)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        conn.close()
        #---------- DONE ----------
##############################################################################
class TXF:
    def __init__(self):
        self.db_name = "TXF"
        self.Main()

    def Main(self):
        #...........................................................................
        wkday_list = self.Update_Date()
        insert_date = []
        ii1_data = []
        ii2_data = []
        ii3_data = []
        
        for wkday in wkday_list:
            print(wkday)
            (data1 , data2 , data3) = self.Data_Fetch(wkday)
            insert_date.append([wkday])
            ii1_data.append(data1)
            ii2_data.append(data2)
            ii3_data.append(data3)

        #print(ii1_data)
        #print(ii2_data)
        #print(ii3_data)
        print("TXF Insert Data To Database")
        DB_Insert(self.db_name , 0 , insert_date)
        DB_Insert(self.db_name , 1 , ii1_data)
        DB_Insert(self.db_name , 2 , ii2_data)
        DB_Insert(self.db_name , 3 , ii3_data)      
        #...........................................................................

    def Update_Date(self):
        #...........................................................................
        if len(Read_DB(self.db_name , 0))==0:
            wkday_list = sorted(set([item[0] for item in Read_DB("TW" , 0) if int(item[3])>=201601]))            
        else:
            wkday_list = sorted(set([item[0] for item in Read_DB("TW" , 0)]))[-1:]
        return(wkday_list)
        #...........................................................................

    def Data_Fetch(self , dates):
        #...........................................................................
        #ceDates = dates.split("-")

        ceDates = dates.replace("-","/")
        
        my_url = "http://www.taifex.com.tw/cht/3/futContractsDate"
        
        my_data = {
            "commodityId" : "MXF",
            "dateaddcnt" : "-1" ,
            "doQuery" : "1" ,
            "goDay" : "",
            "queryDate" : ceDates ,
            "queryType" : "1" ,
            }
        
        time.sleep(5)
        
        r = Url_Fetch(my_url , my_data)
        
        r.encoding = "utf-8"
        
        content = r.text

        pattern = r"""\s*<TR class="12bk">\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">1</div></TD>\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">\s*\s*小型臺指期貨</div>\s*\s*</TD>\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*自營商</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*\s*<TR class="12bk">\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*投信</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*\s*<TR class="12bk">\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*外資</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*"""
	 
        reg = re.compile(pattern)

        match = reg.findall(content)[0]
        
        if not len(match)==18:
            print("[Fetch Fail Alarm]TXF Fetch Counter Is Not Match With 18")
            print("Please Check It !!!")
            input("Please eNter Any To Stop Procedure...")
            sys.exit(0)

        group1 = [dates]
        group2 = [dates]
        group3 = [dates]

        #<<< 自營商 >>>
        for m in match[:6]:
            # print(m , len(m) , type(m))
            for rw in [" " , ">" , "<" , "=" , ","]:
                m = str(m).replace(rw , "")
            group1.append(m)
            
        #<<< 投 信 >>>
        for m in match[6:12]:
            # print(m , len(m) , type(m))
            for rw in [" " , ">" , "<" , "=" , ","]:
                m = str(m).replace(rw , "")
            group2.append(m)  

        #<<< 外 資 >>>
        for m in match[12:]:
            # print(m , len(m) , type(m))
            for rw in [" " , ">" , "<" , "=" , ","]:
                m = str(m).replace(rw , "")
            group3.append(m)

        return(group1 , group2 , group3)
        #...........................................................................
########################################################################################################
class STXF_DB_Create:
    def __init__(self):
        self.DB_Create()

    def DB_Create(self):
        my_db = DB_Path()[4]    #["twse.db", "tx.db" , "txo.db" , "txf.db" ,"stxf.db"]
        #...........................................................................
        # Raw Data DB - Table Lists
        # Table #0 : working_day
        # Table #1 : ii1                     * 自營商
        # Table #2 : ii2                     * 投信
        # Table #3 : ii3                     * 外資
        #...........................................................................
        conn = sqlite3.connect(my_db)
        #...........................................................................
        # <<<<< Table : working day >>>>>
        #...........................................................................
        table0 = "working_day"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table0 + """(
               Date    TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table0  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii1 >>>>>
        # Date
        # Long OI = 多方OI
        # Long OI Money = 多方OI 金額
        # Short OI = 空方OI
        # Short OI Money = 空方OI 金額
        # Diff OI = 多空方OI差
        # Diff OI Money = 多空方OI 金額差
        #...........................................................................
        table1 = "ii1"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table1 + """(
               Date    TEXT    NOT NULL,
               Long_OI  TEXT    NOT NULL,
               Long_OI_Money  TEXT    NOT NULL,
               Short_OI  TEXT    NOT NULL,
               Short_OI_Money  TEXT    NOT NULL,
               Diff_OI  TEXT    NOT NULL,
               Diff_OI_Money  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table1  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii2 >>>>>
        #...........................................................................
        table2 = "ii2"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table2 + """(
               Date    TEXT    NOT NULL,
               Long_OI  TEXT    NOT NULL,
               Long_OI_Money  TEXT    NOT NULL,
               Short_OI  TEXT    NOT NULL,
               Short_OI_Money  TEXT    NOT NULL,
               Diff_OI  TEXT    NOT NULL,
               Diff_OI_Money  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table2  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii3 >>>>>
        #...........................................................................
        table3 = "ii3"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table3 + """(
               Date    TEXT    NOT NULL,
               Long_OI  TEXT    NOT NULL,
               Long_OI_Money  TEXT    NOT NULL,
               Short_OI  TEXT    NOT NULL,
               Short_OI_Money  TEXT    NOT NULL,
               Diff_OI  TEXT    NOT NULL,
               Diff_OI_Money  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table3  + "(Date)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        conn.close()
        #---------- DONE ----------
##############################################################################
class STXF:
    def __init__(self):
        self.db_name = "STXF"
        self.Main()

    def Main(self):
        #...........................................................................
        wkday_list = self.Update_Date()
        insert_date = []
        ii1_data = []
        ii2_data = []
        ii3_data = []
        
        for wkday in wkday_list:
            print(wkday)
            (data1 , data2 , data3) = self.Data_Fetch(wkday)
            insert_date.append([wkday])
            ii1_data.append(data1)
            ii2_data.append(data2)
            ii3_data.append(data3)

        #print(ii1_data)
        #print(ii2_data)
        #print(ii3_data)
        print("Insert Data To Database")        
        DB_Insert(self.db_name , 0 , insert_date)
        DB_Insert(self.db_name , 1 , ii1_data)
        DB_Insert(self.db_name , 2 , ii2_data)
        DB_Insert(self.db_name , 3 , ii3_data)      
        #...........................................................................

    def Update_Date(self):
        #...........................................................................
        if len(Read_DB(self.db_name , 0))==0:
            #Data Start From 2015/12/29
            wkday_list = sorted(set([item[0] for item in Read_DB("TW" , 0) if int(item[3])>=201601]))            
        else:
            wkday_list = sorted(set([item[0] for item in Read_DB("TW" , 0)]))[-1:]
        return(wkday_list)
        #...........................................................................

    def Data_Fetch(self , dates):
        #...........................................................................
        #ceDates = dates.split("-")

        ceDates = dates.replace("-","/")
        
        my_url = "http://www.taifex.com.tw/cht/3/futContractsDate"
        
        my_data = {
            "commodityId" : "MXF",
            "dateaddcnt" : "" ,
            "doQuery" : "1" ,
            "goDay" : "",
            "queryDate" : ceDates ,
            "queryType" : "1" ,
            }
        
        time.sleep(5)
        
        r = Url_Fetch(my_url , my_data)
        
        r.encoding = "utf-8"
        
        content = r.text

        #pattern = r"""\s*<TR class="12bk">\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">1</div></TD>\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">\s*\s*臺股期貨</div>\s*\s*</TD>\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*自營商</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*"""

        #pattern = r"""\s*<TR class="12bk">\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">1</div></TD>\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">\s*\s*臺股期貨</div>\s*\s*</TD>\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*自營商</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*\s*<TR class="12bk">\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*投信</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*\s*<TR class="12bk">\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*外資</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*"""
        pattern = r"""\s*<TR class="12bk">\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">1</div></TD>\s*\s*<TD  rowspan="3" bgcolor='#FFFFF0'><div align="center">\s*\s*小型臺指期貨</div>\s*\s*</TD>\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*自營商</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*\s*<TR class="12bk">\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*投信</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*\s*<TR class="12bk">\s*\s*<TD  bgcolor='#FFFFF0'>\s*\s*<div align="center">\s*\s*外資</div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right">\s*\s*<font color="blue">\s*\s*\S*</font>\s*\s*</div>\s*\s*<div align="right"></div>\s*\s*</TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'>\s*\s*<div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*\S*</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*\S*</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right"><font color="blue">\s*\s*(\S*)</font></div>\s*\s*<div align="right"></div></TD>\s*\s*<TD  align=right nowrap bgcolor='#FFFFF0'><div align="right">\s*\s*(\S*)</div>\s*\s*<div align="right"></div></TD>\s*\s*</TR>\s*"""
        
        reg = re.compile(pattern)

        match = reg.findall(content)[0]
        
        if not len(match)==18:
            print("[Fetch Fail Alarm]STXF Fetch Counter Is Not Match With 18")
            print("Please Check It !!!")
            input("Please eNter Any To Stop Procedure...")
            sys.exit(0)

        group1 = [dates]
        group2 = [dates]
        group3 = [dates]

        #<<< 自營商 >>>
        for m in match[:6]:
            # print(m , len(m) , type(m))
            for rw in [" " , ">" , "<" , "=" , ","]:
                m = str(m).replace(rw , "")
            group1.append(m)
            
        #<<< 投 信 >>>
        for m in match[6:12]:
            # print(m , len(m) , type(m))
            for rw in [" " , ">" , "<" , "=" , ","]:
                m = str(m).replace(rw , "")
            group2.append(m)  

        #<<< 外 資 >>>
        for m in match[12:]:
            # print(m , len(m) , type(m))
            for rw in [" " , ">" , "<" , "=" , ","]:
                m = str(m).replace(rw , "")
            group3.append(m)

        return(group1 , group2 , group3)
        #...........................................................................
    
########################################################################################################
class TXOP_DB_Create:
    def __init__(self):
        self.DB_Create()

    def DB_Create(self):
        my_db = DB_Path()[5]    #["twse.db", "tx.db" , "txo.db" , "txf.db" ,"stxf.db" , "txop.db"]
        #...........................................................................
        # Raw Data DB - Table Lists
        # Table #1 : working_day
        # Table #2 : ii1                     * 自營商
        # Table #3 : ii2                     * 投信
        # Table #4 : ii3                     * 外資
        #...........................................................................
        conn = sqlite3.connect(my_db)
        #...........................................................................
        # <<<<< Table : working day >>>>>
        #...........................................................................
        table0 = "working_day"        
        cmd = "CREATE TABLE IF NOT EXISTS " + table0 + """(
               Date    TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table0  + "(Date)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        # <<<<< Table1 : ii1 >>>>>
        # Date
        # Call_Buy_OI
        # Call_Buy_Valume
        # Call_Sell_OI
        # Call_Sell_Valume
        # Call_Diff_OI
        # Call_Diff_Valume
        # Put_Buy_OI
        # Put_Buy_Valume
        # Put_Sell_OI
        # Put_Sell_Valume
        # Put_Diff_OI
        # Put_Diff_Valume
        #...........................................................................
        table1 = "ii1"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table1 + """(
               Date                         TEXT    NOT NULL,
               Call_Buy_OI            TEXT    NOT NULL,
               Call_Buy_Valume  TEXT    NOT NULL,
               Call_Sell_OI            TEXT    NOT NULL,
               Call_Sell_Valume  TEXT    NOT NULL,
               Call_Diff_OI            TEXT    NOT NULL,
               Call_Diff_Valume  TEXT    NOT NULL,
               Put_Buy_OI            TEXT    NOT NULL,
               Put_Buy_Valume  TEXT    NOT NULL,
               Put_Sell_OI            TEXT    NOT NULL,
               Put_Sell_Valume  TEXT    NOT NULL,
               Put_Diff_OI            TEXT    NOT NULL,
               Put_Diff_Valume  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table1  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii2 >>>>>
        #...........................................................................
        table2 = "ii2"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table2 + """(
               Date                         TEXT    NOT NULL,
               Call_Buy_OI            TEXT    NOT NULL,
               Call_Buy_Valume  TEXT    NOT NULL,
               Call_Sell_OI            TEXT    NOT NULL,
               Call_Sell_Valume  TEXT    NOT NULL,
               Call_Diff_OI            TEXT    NOT NULL,
               Call_Diff_Valume  TEXT    NOT NULL,
               Put_Buy_OI            TEXT    NOT NULL,
               Put_Buy_Valume  TEXT    NOT NULL,
               Put_Sell_OI            TEXT    NOT NULL,
               Put_Sell_Valume  TEXT    NOT NULL,
               Put_Diff_OI            TEXT    NOT NULL,
               Put_Diff_Valume  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table2  + "(Date)"
        conn.execute(cmd)
        #...........................................................................
        
        #...........................................................................
        # <<<<< Table : ii3 >>>>>
        #...........................................................................
        table3 = "ii3"         
        cmd = "CREATE TABLE IF NOT EXISTS " + table3 + """(
               Date                         TEXT    NOT NULL,
               Call_Buy_OI            TEXT    NOT NULL,
               Call_Buy_Valume  TEXT    NOT NULL,
               Call_Sell_OI            TEXT    NOT NULL,
               Call_Sell_Valume  TEXT    NOT NULL,
               Call_Diff_OI            TEXT    NOT NULL,
               Call_Diff_Valume  TEXT    NOT NULL,
               Put_Buy_OI            TEXT    NOT NULL,
               Put_Buy_Valume  TEXT    NOT NULL,
               Put_Sell_OI            TEXT    NOT NULL,
               Put_Sell_Valume  TEXT    NOT NULL,
               Put_Diff_OI            TEXT    NOT NULL,
               Put_Diff_Valume  TEXT    NOT NULL,
               UNIQUE(Date))"""
        conn.execute(cmd)
        
        #Index - Date
        cmd = "CREATE INDEX IF NOT EXISTS Idx_Date ON " + table3  + "(Date)"
        conn.execute(cmd)
        #...........................................................................

        #...........................................................................
        conn.close()
        #---------- DONE ----------
##############################################################################
class TXOP:
    def __init__(self):
        self.db_name = "TXOP"
        self.Main()

    def Main(self):
        #...........................................................................
        wkday_list = self.Update_Date()
        insert_date = []
        ii1_data = []
        ii2_data = []
        ii3_data = []
        
        for wkday in wkday_list:
            print(wkday)
            (data1 , data2 , data3) = self.Data_Fetch(wkday)
            insert_date.append([wkday])
            ii1_data.append(data1)
            ii2_data.append(data2)
            ii3_data.append(data3)

        #print(ii1_data)
        #print(ii2_data)
        #print(ii3_data)
        print("Insert Data To Database")
        
        DB_Insert(self.db_name , 0 , insert_date)
        DB_Insert(self.db_name , 1 , ii1_data)
        DB_Insert(self.db_name , 2 , ii2_data)
        DB_Insert(self.db_name , 3 , ii3_data)      
        #...........................................................................

    def Update_Date(self):
        #...........................................................................
        if len(Read_DB(self.db_name , 0))==0:
            wkday_list = sorted(set([item[0] for item in Read_DB("TW" , 0) if int(item[3])>=201601]))            
        else:
            wkday_list = sorted(set([item[0] for item in Read_DB("TW" , 0)]))[-1:]
        return(wkday_list)
        #...........................................................................

    def Data_Fetch(self , dates):
        #...........................................................................
        #ceDates = dates.split("-")

        ceDates = dates.replace("-","/")
        
        my_url = "http://www.taifex.com.tw/cht/3/callsAndPutsDate"
        
        my_data = {
            "commodityId" : "TXO",
            "dateaddcnt" : "" ,
            "doQuery" : "1" ,
            "goDay" : "",
            "queryDate" : ceDates ,
            "queryType" : "1" ,
            }
        
        time.sleep(5)
        
        r = Url_Fetch(my_url , my_data)
        
        r.encoding = "utf-8"
        
        content = r.text

        pattern = r"""\s*<TD align=right nowrap bgcolor='#\w+F0'><font color="blue">\s*\S+</font></TD>\s*\s*<TD height="15" align=right nowrap bgcolor='#\w+F0'>\s*\s*<div align="right"></div>\s*\s*<div align="right">\s*\S+</TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'><font color="blue">\s*\S+</font></TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'>\s*\S+</TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'><font color="blue">\s*\S+</font></TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'>\s*\S+</TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'><font color="blue">\s*\s*(\S+)</font></TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'>\s*\s*(\S+)</TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'><font color="blue">\s*\s*(\S+)</font></TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'>\s*\s*(\S+)</TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'><font color="blue">\s*\s*(\S+)</font></TD>\s*\s*<TD align=right nowrap bgcolor='#\w+F0'>\s*\s*(\S+)</TD>\s*\s*</TR>"""
        
        reg = re.compile(pattern)

        #match = reg.findall(content)[0]
        match = reg.findall(content)
                
        group1 = [dates]
        group2 = [dates]
        group3 = [dates]

        if not len(match)==6:
            print("[Fetch Fail Alarm]TXOP Fetch Counter Is Not Match With 6")
            print("Please Check It !!!")
            input("Please eNter Any To Stop Procedure...")
            sys.exit(0)

        for index ,raw in enumerate(match):
            if not len(raw)==6:
                print("[Fetch Fail Alarm]TXOP Fetch Raw Data Is Not Match With 6")
                print("Please Check It !!!")
                input("Please eNter Any To Stop Procedure...")
                sys.exit(0)
                
            #<<< 自營商 >>>
            if index==0 or index==3:
                for m in match[index]:
                    group1.append(m)
                    
            elif index==1 or index==4:
                #<<< 投 信 >>>
                for m in match[index]:
                    group2.append(m)
                    
            elif index==2 or index==5:
                #<<< 外 資 >>>
                for m in match[index]:
                    group3.append(m)
                    
        return(group1 , group2 , group3)
        #...........................................................................
########################################################################################################
    
class Daily_Report:
    def __init__(self):
        self.Main()

    def Main(self):
        #...........................................................................
        self.TW_CSV_Report_Daily()
        self.TX_CSV_Report_Daily()
        self.TW_Xlsx_Report_Daily()
        self.TXO_Xlsx_OI_Report_Daily()        
        self.TXO_Month_OP_Report_Daily()
        #...........................................................................

    def TXO_Month_OP_Report_Daily(self):
        #...........................................................................
        mon_c_list = []
        db_date = sorted(set(Read_DB("TXO" , 0)))
        for no , item in enumerate(db_date):
            if not item[1][-2]=='W':
                index_price = int(item[-1])
                c = int(index_price / 100)
                mon_c_list.append([item[0] , item[1] , item[2] , (c+1)*100 , c*100])

        #
        mon_c_list = mon_c_list[-1]
        Pre_TXO = mon_c_list[1]
        TXO_Year = int(Pre_TXO[:4])
        TXO_Mon = int(Pre_TXO[-2:])
        
        if (TXO_Mon+1)>=13:
            TXO = str(TXO_Year+1)+str(1).zfill(2)
        else:
            TXO = str(TXO_Year)+str(TXO_Mon+1).zfill(2)

        s_date  = mon_c_list[0]
        e_date = time.strftime('%Y-%m-%d', time.localtime())
        
        tw_querty = self.Mon_Read_DB("TW" , 1 , s_date , e_date)
        tx_querty = self.Mon_Read_DB("TX" , 3 , s_date , e_date)
        #print(TXO , mon_c_list[-1] , s_date , e_date)
        self.Mon_OP_Report(TXO , tw_querty , tx_querty )
        #...........................................................................

    def Mon_Read_DB(self , DB_Type , Table_No , S_Date , E_Date):
        #...........................................................................
        if DB_Type.upper()=="TW":
            my_db = DB_Path()[0]    #["twse.db", "tx.db"]
            #....................................................
            # Table #0 : working_day
            # Table #1 : daily
            # Table #2 : weekly
            # Table #3 : monthly
            #....................................................
            if Table_No<=-1 or Table_No>3:
                print("TW DB Table List Error!!!Please Double Confirm")
                return(None)
            else:
                myTable = {0:"working_day" ,  1:"daily" , 2:"weekly" , 3:"monthly"}
                
        elif DB_Type.upper()=="TX":
            #...........................................................................
            # Table #0 : source
            # Table #1 : working_day
            # Table #2 : contract_list  #契約列表
            # Table #3 : daily
            # Table #4 : weekly
            # Table #5 : monthly
            #...........................................................................
            my_db = DB_Path()[1]    #["twse.db", "tx.db"]
            
            if Table_No<=-1 or Table_No>6:
                print("TX DB Table List Error!!!Please Double Confirm")
                return(None)
            else:
                myTable = {0 : "source" , 1 : "working_day" ,  2 : "contract_list" , 3 : "daily" , 4 : "weekly" , 5 : "monthly"}               
        else:
            print("DB Name Error!!!Please Double Confirm")
            return(None)

        #DB SETUP
        my_table = myTable[Table_No]
        
        #
        #print(my_db , my_table)
        conn = sqlite3.connect(my_db)
        cur = conn.cursor()
        cmd = "SELECT * FROM " + my_table  + " WHERE Date>='" + S_Date + "' AND Date<='" + E_Date + "'"
        cur.execute(cmd)
        querty = cur.fetchall()
        conn.close()
        return(querty)
        #...........................................................................


    def Mon_OP_Report(self , TXO , tw_querty , tx_querty):
        wb = Workbook()
        ws = wb.active
        ws.title = TXO
        myws = ws

        #myTitle =["日期" , "加權收盤" ,  "期貨收盤" ,"價平履約價" , "CALL收盤價","PUT收盤價","C OI" , "P OI" , "C+P價格"]
        Title1 = ["日期" , "開盤價" , "最高價" , "最低價" , "收盤價" , "成交量(億)"]
        
        for i in range(len(Title1)):
        #row = 1 ,2 , 3 , 4 .....
        #column = A , B , C , D ...
            myws.cell(row=1, column=i+1).value = Title1[i]
            myws.cell(row=1, column=i+1).alignment = Alignment(horizontal="center", vertical="center")
            myws.cell(row=1, column=i+1).font = Font(name='Arial' , size=12)

            myws.cell(row=1, column=i+8).value = Title1[i]
            myws.cell(row=1, column=i+8).alignment = Alignment(horizontal="center", vertical="center")
            myws.cell(row=1, column=i+8).font = Font(name='Arial' , size=12)  

        #TWSE
        data_col = [0,4,5,6,7,8]
        for index , row in enumerate(tw_querty , start=2):
            for i in range(len(data_col)):
                if i>=1:
                    myws.cell(row=index, column=i+1).value = float(str(row[data_col[i]]).replace(",",""))
                else:
                    myws.cell(row=index, column=i+1).value = row[data_col[i]]

        #TX
        for index , row in enumerate(tx_querty , start=2):
            for i in range(len(data_col)):
                myws.cell(row=index, column=i+8).value = row[data_col[i]]
                    
        len1 = len(tw_querty)
        len2 = len(tx_querty)

        
        #---------------------------------------------------------------------------------
        chart_h = 8
        chart_w = 18
        #---------------------------------------------------------------------------------          
        ws1 = wb.create_sheet("Chart")
        c1 = StockChart()
        s_row = 2
        e_row = len1+1
        #col : A , B , C ...
        #Row : 1 , 2 , 3 ...
        labels = Reference(ws , min_col=1, min_row=s_row, max_row=e_row)
        data = Reference(ws , min_col=2, max_col=5, min_row=s_row, max_row=e_row)
        c1.add_data(data, titles_from_data=False)
        c1.set_categories(labels)
        for s in c1.series:
            s.graphicalProperties.line.noFill = True
        c1.hiLowLines = ChartLines()
        c1.upDownBars = UpDownBars()
        #c.title = "Open-high-low-close"

        # Excel is broken and needs a cache of values in order to display hiLoLines :-/
        from openpyxl.chart.data_source import NumData, NumVal
        pts = [NumVal(idx=i) for i in range(len1-1)]
        cache = NumData(pt=pts)
            
        # add dummy cache
        c1.series[-1].val.numRef.numCache = cache
            
        # Create bar chart for volume
        bar1 = BarChart()
        data =  Reference(ws, min_col=6, min_row=s_row, max_row=e_row)
        bar1.add_data(data, titles_from_data=False)
        bar1.set_categories(labels)
        #---------------------------------------------------------------------------------
        from copy import deepcopy

        b1 = deepcopy(bar1)
        c2 = deepcopy(c1)
        c2.y_axis.majorGridlines = None
        #c1.y_axis.title = "Price"
        b1.y_axis.axId = 20
        b1.z_axis = c1.y_axis
        b1.y_axis.crosses = "max"
        b1 += c2
        b1.height = chart_h
        b1.width = chart_w
        ws1.add_chart(b1, "A1")
        #---------------------------------------------------------------------------------
        #---------------------------------------------------------------------------------
        c3 = StockChart()
        s_row = 2
        e_row = len2+1
        #col : A , B , C ...
        #Row : 1 , 2 , 3 ...
        labels = Reference(ws , min_col=8, min_row=s_row, max_row=e_row)
        data = Reference(ws , min_col=9, max_col=12, min_row=s_row, max_row=e_row)
        c3.add_data(data, titles_from_data=False)
        c3.set_categories(labels)
        for s in c3.series:
            s.graphicalProperties.line.noFill = True
        c3.hiLowLines = ChartLines()
        c3.upDownBars = UpDownBars()
        #c.title = "Open-high-low-close"

        # Excel is broken and needs a cache of values in order to display hiLoLines :-/
        from openpyxl.chart.data_source import NumData, NumVal
        pts = [NumVal(idx=i) for i in range(len2-1)]
        cache = NumData(pt=pts)
            
        # add dummy cache
        c3.series[-1].val.numRef.numCache = cache
            
        # Create bar chart for volume
        bar2 = BarChart()
        data =  Reference(ws, min_col=13, min_row=s_row, max_row=e_row)
        bar2.add_data(data, titles_from_data=False)
        bar2.set_categories(labels)
        #---------------------------------------------------------------------------------
        from copy import deepcopy

        b2 = deepcopy(bar2)
        c4 = deepcopy(c3)
        c4.y_axis.majorGridlines = None
        #c1.y_axis.title = "Price"
        b2.y_axis.axId = 20
        b2.z_axis = c3.y_axis
        b2.y_axis.crosses = "max"
        b2 += c4
        b2.height = chart_h
        b2.width = chart_w
        ws1.add_chart(b2, "L1")
        #---------------------------------------------------------------------------------
        #---------------------------------------------------------------------------------
        rpt_dir = Daily_Report_Dir()
        xl_file = os.path.join(rpt_dir  , "Month_OP_Daily_Report-" + time.strftime("%Y%m%d", time.localtime()) + '.xlsx')
        wb.save(xl_file)
        #...........................................................................


    def TXO_Xlsx_OI_Report_Daily(self):
        #...........................................................................
        db_date = sorted(set(Read_DB("TXO" , 1)))[-1][0]

        csv_name = db_date.replace("-", "") + "_txo.csv"
            
        csv_file = Dir_Chk()[-1] + "\\" + db_date[:4] + "\\" + db_date[:4] + db_date[5:7] + "\\" + csv_name

        if os.path.isfile(csv_file):            
            with open(csv_file , 'r') as fp:
                content = csv.reader(fp)
                exp_list = [ row[1] for row in content if len(row)>0]
                exp_list = sorted(set(exp_list))
                    
            #exp_wks = [ row for row in exp_list if not row[-1]==")" and len(row)>6]
            exps = [ row for row in exp_list if not row[-1]==")" and len(row)>=6]
            
        else:
            exps = []

        if len(exps)>0:
            wb = Workbook()
            
            for index , exp in  enumerate(exps):
                my_db = os.path.join(Dir_Chk()[-2] , "TXO" + exp[:4] +".db.")
                
                my_table = "OP_" + exp
                
                conn = sqlite3.connect(my_db)
                
                cur = conn.cursor()
                
                cmd = "SELECT Strike_Price FROM " + my_table + " WHERE Date='" + db_date + "' AND BS_Type='買權'"
                cur.execute(cmd)
                target_price = sorted(set([int(row[0]) for row in cur.fetchall()]))
                
                for op_type in ["買權" , "賣權"]:
                    cmd = "SELECT * FROM " + my_table + " WHERE Date='" + db_date + "' AND BS_Type='" + op_type + "'"
                    cur.execute(cmd)
                    querty = cur.fetchall()
                    if op_type=="買權":
                        call_data = querty
                    else:
                        put_data = querty
                conn.close()

                #Excel Content
                if index==0:
                    ws = wb.active
                    ws.title = "TXO" + exp
                    myws = ws
                else:
                    ws1 = wb.create_sheet("TXO" + exp)
                    myws = ws1

                num = len(target_price)
        
                #titleA =[[db_date , "履約價" ,  db_date]]
                titleA =[["買權" , "履約價" ,  "賣權"]]
                for index , row in enumerate(titleA , start=1):
                    #1 ['s14495', '6206', 'Wder']
                    #row = 1 ,2 , 3 , 4 .....
                    #column = A , B , C , D ...
                    for i in range(len(row)):
                        myws.cell(row=index, column=i+1).value = row[i]
                        myws.cell(row=index, column=i+1).alignment = Alignment(horizontal="center", vertical="center")
                        myws.cell(row=index, column=i+1).font = Font(name='Arial' , size=12 , color='FFFFFF')
                        myws.cell(row=index, column=i+1).fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('0000FF'))

                for index , row in enumerate(target_price , start=2):
                        myws.cell(row=index, column=2).value = row
                        myws.cell(row=index, column=2).alignment = Alignment(horizontal="center", vertical="center")
                        myws.cell(row=index, column=2).font = Font(name='Arial' , size=12 , color='FFFFFF')
                        myws.cell(row=index, column=2).fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('550088'))

                for index , row1 in enumerate(target_price  , start=2):
                    for row2 in call_data:
                        if row2[1]==str(row1):
                            myws.cell(row=index, column=1).value = row2[-1]
                            
                    for row2 in put_data:
                        if row2[1]==str(row1):
                            myws.cell(row=index, column=3).value = row2[-1]
                            
                c1 = BarChart()
                c1.type = "col"
                c1.style = 10

                data1 = Reference(myws, min_col=1, min_row=1, max_row=num+1)
                data2 = Reference(myws, min_col=3, min_row=1, max_row=num+1)
                cats = Reference(myws, min_col=2, min_row=2, max_row=num+1)
                c1.add_data(data1, titles_from_data=True)
                c1.add_data(data2, titles_from_data=True)
                c1.set_categories(cats)
                c1.shape = 4
                c1.width = 27
                c1.height = 15
                myws.add_chart(c1, "D1")
            #........................................................................................................................................................
            rpt_dir = Daily_Report_Dir()
            xl_file = os.path.join(rpt_dir  , "TXO_Daily_OI_Report-" + time.strftime("%Y%m%d", time.localtime()) + '.xlsx')
            wb.save(xl_file)
            #...........................................................................


    def TW_Xlsx_Report_Daily(self):
        #...........................................................................
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        #ws = wb.create_sheet("Data")
        #ft = Font(color=colors.RED)
        #for i in range(1,11):
        #   ws["A" + str(i)].value = i
        #    ws["A" + str(i)].font = Font(color=colors.RED , name='Arial' , size=i)

        #Create Title
        #........................................................................................................................................................
        #V1.1.11
        tw_title = [["日期" , "開盤價" , "最高價" , "最低價" , "收盤價" , "成交量(億)" ,
                    "3MA" , "5MA" , "10MA" , "20MA" , "60MA" , "K/D(9,3,3)"]]
        
        kd_title = [["K(%)" , "D(%)" , "Status"]]

        remark_title = [["備註1" , "備註2"]]

        for index , row in enumerate(tw_title , start=1):
            #1 ['s14495', '6206', 'Wder']
            #row = 1 ,2 , 3 , 4 .....
            #column = A , B , C , D ...
            for i in range(len(row)):
                ws.cell(row=index, column=i+1).value = row[i]
                ws.cell(row=index, column=i+1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=index, column=i+1).font = Font(name='Arial' , size=12 , color='FFFFFF')
                
        for index , row in enumerate(kd_title , start=2):
            for i in range(len(row)):
                ws.cell(row=index, column=i+12).value = row[i]
                ws.cell(row=index, column=i+12).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=index, column=i+12).font = Font(name='Arial' , size=12 , color='FFFFFF')
                
        for index , row in enumerate(remark_title , start=1):
            for i in range(len(row)):
                ws.cell(row=index, column=i+15).value = row[i]
                ws.cell(row=index, column=i+15).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=index, column=i+15).font = Font(name='Arial' , size=12 , color='FFFFFF')
    
        #........................................................................................................................................................
        #<<< Cell Merge >>>
        #........................................................................................................................................................
        merge_cells = ["A" , "B" , "C" , "D" , "E" , "F" , "G" , "H" , "I", "J" , "K", "L"]

        for col in merge_cells[:-1]:
            cell_range =col +"1:" + col +"2"
            ws.merge_cells(cell_range)
                
        #For KD
        ws.merge_cells("L1:N1")
        
        for col in ["O"]:
            cell_range =col +"1:" + col +"2"
            ws.merge_cells(cell_range)
            
        ws.freeze_panes = ws["A3"]    
        #........................................................................................................................................................
        #<<< Cell Background Color >>>
            #........................................................................................................................................................
        for col in ["A"]:
            ws[col + "1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('000000'))
            
        for col in ["B" , "C" , "D" , "E"]:
            ws[col + "1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('0000FF'))
            
        for col in ["F"]:
            ws[col + "1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('880000'))
            
        for col in ["G" , "H" , "I" , "J" , "K"]:
            ws[col + "1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('000088'))

        for col in ["L"]:
            ws[col + "1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('3A0088'))
            
        for col in ["L" , "M" , "N"]:
            ws[col + "2"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('3A0088'))
    
        for col in ["O"]:
            ws[col + "1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=colors.Color('A42D00'))
            
        #........................................................................................................................................................
        # TWII Row Data
        #........................................................................................................................................................
        if True:
            #ws['A1'].value = "weder"
            sql_querty = Read_DB("TW" , 1)

            date_lists = [[item[0]] for item in sorted(set(sql_querty))]

            row_data = [[float(item[4].replace(",","")) , float(item[5].replace(",","")) , float(item[6].replace(",","")) , float(item[7].replace(",",""))]for item in sorted(set(sql_querty))]
            
            vol_data = [[item[8]] for item in sorted(set(sql_querty))]

            close = [float(item[7].replace(",","")) for item in sorted(set(sql_querty))]
                    
            ma = MA(close) #[ma3 , ma5 , ma10 , ma20 , ma60] = [10743.19, 10743.75, 10737.19, 10579.64, 10517.46]

            kdj = KDJ(row_data)

            #Ver 1.1.11
            (mv_line , diff_line) = Move_Line(close)

            num = len(date_lists)

            for i in range(num):
                index = i +3
                ws.cell(row=index, column=1).value = date_lists[i][0]

                #OP / HP / LP / CP
                for j in range(4):
                    ws.cell(row=index, column=j+2).value = row_data[i][j]

                #Vol
                ws.cell(row=index, column=6).value = round(vol_data[i][0],2)
                ws.cell(row=index, column=6).font = Font(bold=True)

                #3MA / 5MA / 10MA / 20MA / 60MA
                for j in range(5):
                    ws.cell(row=index, column=j+7).value = int(round(ma[i][j],0))
                    
                #K/D
                for j in range(2):
                    ws.cell(row=index, column=j+12).value = round(kdj[i][j],2)
                    
        #........................................................................................................................................................
        # Cell Style
        #........................................................................................................................................................
            for i in range(num):
                for j in range(1,17):
                    ws.cell(row=i+3, column=j).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=i+3, column=j).font = Font(name='Arial' , size=12)

        #........................................................................................................................................................
        # Cell Style
        #........................................................................................................................................................
        for i in range(59 , num):
            index = 3+i
            (lop , lhp , llp, lcp, lk , ld) = ws["B" + str(index-1)].value , ws["C" + str(index-1)].value , ws["D" + str(index-1)].value , ws["E" + str(index-1)].value , ws["L" + str(index-1)].value, ws["M" + str(index-1)].value
            (op , hp , lp, cp, vol , ma3 , ma5 , ma10 , ma20 , ma60 , k , d) = (ws["B" + str(index)].value , ws["C" + str(index)].value , ws["D" + str(index)].value , ws["E" + str(index)].value , ws["F" + str(index)].value , ws["G" + str(index)].value , ws["H" + str(index)].value , ws["I" + str(index)].value , ws["J" + str(index)].value , ws["K" + str(index)].value , ws["L" + str(index)].value , ws["M" + str(index)].value)
            if cp<ma3:ws["G" +str(index)].font = Font(color='227700' , name='Arial' , size=12)
            else:ws["G" +str(index)].font = Font(color='FF0000' , name='Arial' , size=12)
                
            if cp<ma5:ws["H" +str(index)].font = Font(color='227700' , name='Arial' , size=12)
            else:ws["H" +str(index)].font = Font(color='FF0000' , name='Arial' , size=12)

            if cp<ma10:ws["I" +str(index)].font = Font(color='227700' , name='Arial' , size=12)
            else:ws["I" +str(index)].font = Font(color='FF0000' , name='Arial' , size=12)
            
            if cp<ma20:ws["J" +str(index)].font = Font(color='227700' , name='Arial' , size=12)
            else:ws["J" +str(index)].font = Font(color='FF0000' , name='Arial' , size=12)

            if cp<ma60:ws["K" +str(index)].font = Font(color='227700' , name='Arial' , size=12)
            else:ws["K" +str(index)].font = Font(color='FF0000' , name='Arial' , size=12)

            if lk>ld and k<d:
                ws["N" +str(index)].value="死叉"
                ws["N" +str(index)].font = Font(color='227700' , name='Arial' , size=12 , bold=True)
                
            elif lk<ld and k>d:
                ws["N" +str(index)].value="金叉"
                ws["N" +str(index)].font = Font(color='FF0000' , name='Arial' , size=12 , bold=True)
                
            #Colume P
            #if vol>1200 and vol<1500:
            #    ws["O" +str(index)].value="成交量突破1200億"
            #    ws["O" +str(index)].font = Font(color='0000FF' , name='Arial' , size=14 , bold=True)
            if vol<800:
                ws["O" +str(index)].value="窒息量"
                ws["O" +str(index)].font = Font(color='0000FF' , name='Arial' , size=12 , bold=True)
                
            elif vol>=1500:
                ws["O" +str(index)].value="爆大量"
                ws["O" +str(index)].font = Font(color='FF0000' , name='Arial' , size=12 , bold=True)

                    
        #........................................................................................................................................................
        # Chart
        #........................................................................................................................................................  
        ws1 = wb.create_sheet("Chart")
        c1 = StockChart()
        s_row = num+2-90
        e_row = num+2
        #col : A , B , C ...
        #Row : 1 , 2 , 3 ...
        labels = Reference(ws , min_col=1, min_row=s_row, max_row=e_row)
        data = Reference(ws , min_col=2, max_col=5, min_row=s_row, max_row=e_row)
        c1.add_data(data, titles_from_data=False)
        c1.set_categories(labels)
        for s in c1.series:
            s.graphicalProperties.line.noFill = True
        c1.hiLowLines = ChartLines()
        c1.upDownBars = UpDownBars()
        #c.title = "Open-high-low-close"

        # Excel is broken and needs a cache of values in order to display hiLoLines :-/
        from openpyxl.chart.data_source import NumData, NumVal
        pts = [NumVal(idx=i) for i in range(num - 1)]
        cache = NumData(pt=pts)
        
        # add dummy cache
        c1.series[-1].val.numRef.numCache = cache
        
        # Create bar chart for volume
        bar1 = BarChart()
        data =  Reference(ws, min_col=6, min_row=s_row, max_row=e_row)
        bar1.add_data(data, titles_from_data=False)
        bar1.set_categories(labels)
        #---------------------------------------------------------------------------------
        from copy import deepcopy

        b1 = deepcopy(bar1)
        c2 = deepcopy(c1)
        c2.y_axis.majorGridlines = None
        #c1.y_axis.title = "Price"
        b1.y_axis.axId = 20
        b1.z_axis = c1.y_axis
        b1.y_axis.crosses = "max"
        b1 += c2
        b1.height = 10
        b1.width = 20.5

        ws1.add_chart(b1, "A1")
        #---------------------------------------------------------------------------------
        c2 = LineChart()
        c2.title = "KD"
        c2.style = 12
        #c2.y_axis.title = "Size"
        #c2.y_axis.crossAx = 500
        #c2.x_axis = DateAxis(crossAx=100)
        #c2.x_axis.number_format = 'd-mmm'
        #c2.x_axis.majorTimeUnit = "days"
        #c2.x_axis.title = "Date"    
        data1 = Reference(ws, min_col=12, min_row=s_row, max_row=e_row)
        data2 = Reference(ws, min_col=13, min_row=s_row, max_row=e_row)
        c2.add_data(data1, titles_from_data= False)
        c2.add_data(data2, titles_from_data= False)
        c2.set_categories(labels)
        c2.height = 10
        c2.width = 20.5

        ws1.add_chart(c2, "A19")

        #........................................................................................................................................................  
        rpt_dir = Daily_Report_Dir()
        xl_file = os.path.join(rpt_dir  , "TW_Daily_Report-" + time.strftime("%Y%m%d", time.localtime()) + '.xlsx')
        wb.save(xl_file)
        #...........................................................................


    def TX_CSV_Report_Daily(self):
        #...........................................................................
        tx_daily_data = Read_DB("TX" , 3)

        datas = tx_daily_data[-90:]
        
        content = ""
        
        #DB
        my_db = DB_Path()[1]
        my_table = "source"
        conn = sqlite3.connect(my_db)
        cur = conn.cursor()
        try:
            for data in datas:
                #print(data)
                for i , item in enumerate(data):
                    if i==0:
                        myDate = item
                        content+= "|" + item.center(12)+"|"
                    elif i>3 and i<8:
                        content+= str(int(item)).center(10)+"|"
                    elif i==8:
                        content+= str(int(item)).rjust(9)+" |"
                    if i==9:
                        content+= str(int(item)).rjust(12)+" | "

                cmd = "SELECT * FROM " + my_table + " WHERE Date='" + myDate +"'"
                cur.execute(cmd)
                querty = cur.fetchall()

                v0 = querty[0][-2]
                oi0 = querty[0][-1]   
                v1 = querty[1][-2]
                oi1 = querty[1][-1]

                if str(v0)=='-':v0=0
                if str(oi0)=='-':oi=0
                if str(v1)=='-':v1=0
                if str(oi1)=='-':oi1=0

                vr0 = round(v0*100/(v0+v1) , 2)
                vr1 = round(v1*100/(v0+v1) , 2)

                oir0 = round(oi0*100/(oi0+oi1) , 2)
                oir1 = round(oi1*100/(oi0+oi1) , 2)

                #..............................................................................................................
                content += str(vr0).ljust(5,"0")+ " | "

                if str(vr1)[-2]==".":
                    content += (str(vr1)+"0").rjust(5)+ " |  "
                else:
                    content +=str(vr1).rjust(5)+ " |  "
                #..............................................................................................................
                content += str(oir0).ljust(5,"0")+ " | "
                
                if str(oir1)[-2]==".":
                    content += (str(oir1)+"0").rjust(5)+ " |  "
                else:
                    content +=str(oir1).rjust(5)+ " |  "
                #..............................................................................................................
                content +="\n"
                
        except:
            print("Date#" , data[0] , " - TX Data Error!")
            print("Procedure Stop and Please Confrim Data Again")
            input("Please Enter Any To Leave!")
            sys.exit(0)
        conn.close()

        return(self.CSV_Daily_Report("TX" , myDate , content))
        #---------- DONE ----------
        #...........................................................................

    def TW_CSV_Report_Daily(self):
        #...........................................................................
        tw_daily_data = Read_DB("TW" , 1)
        
        datas = tw_daily_data[-90:]
        
        content = ""
        for data in datas:
            for i , item in enumerate(data):
                if i==0:
                    myDate = item
                    content+= "|" + item.center(12)+"|"
                elif i>3 and i<8:
                    content+= item.center(10)+"|"
                elif i==8:
                    if str(round(float(item),2))[-3]==".":
                        content+= str(round(float(item),2)).rjust(13)+" |"
                    else:
                        content+= str(round(float(item),2)).rjust(12)+"0 |"
                elif i>8:
                    content+= str(item).center(13)+"|"

                if i==10:
                    content +="\n"
                    
        return(self.CSV_Daily_Report("TW" , myDate , content))
        #---------- DONE ----------
        #...........................................................................

                
    def CSV_Daily_Report(self , data_type , data_date , content):
        #---------------------------------------------------------------------------------------------------------------------------------
        dates = data_date.split("-")

        file_content = ""

        report_path = Daily_Report_Dir()
                
        if data_type=="TW":
            title = """+---------------------------------------------------------------------------------------------------+
+                  台          灣          加          權          指          數                   +
+------------+----------+----------+----------+----------+--------------+-------------+-------------+
|   日  期   | 開 盤 價 | 最 高 價 | 最 低 價 | 收 盤 價 | 成 交 量(億) | 成 交 股 數 | 成 交 筆 數 |
+------------+----------+----------+----------+----------+--------------+-------------+-------------+
"""
            end_line = "+------------+----------+----------+----------+----------+--------------+-------------+-------------+"

            file_name = "台灣加權指數日報表-" + "".join(dates) + ".asc"

            file_path = os.path.join(report_path , file_name)
            
        elif data_type=="TX":
            title = """+------------------------------------------------------------------------------------------------------------------+
+                 台         灣         期         貨         近         月         指         數                  +
+------------+----------+----------+----------+----------+----------+-------------+---------------+----------------+
|            |          |          |          |          |          |             |  成 交 量 比  | 未 平 倉 量 比 |
|   日  期   | 開 盤 價 | 最 高 價 | 最 低 價 | 收 盤 價 | 成 交 量 | 未 平 倉 量 |-------+-------|--------+-------+
|            |          |          |          |          |          |             | 近 月 | 遠 月 |  近 月 | 遠 月 |
+------------+----------+----------+----------+----------+----------+-------------+---------------+----------------+
"""
            end_line = "+------------+----------+----------+----------+----------+----------+-------------+---------------+----------------+"
            
            file_name = "台灣期貨近月指數日報表-" + "".join(dates) + ".asc"

            file_path = os.path.join(report_path , file_name)

        #.....
        file_content = title + content + end_line

        fp = open(file_path , "w" , encoding='utf8')
        
        fp.write(file_content)
        
        fp.close()

        return(file_path)
        #---------- DONE ----------
        #---------------------------------------------------------------------------------------------------------------------------------
########################################################################################################

class GMail:
    def __init__(self):
        self.Main()

    def Main(self):
        #------------------------------------------------------------------
        rpt_dir = Daily_Report_Dir()

        attacheds = [os.path.join(rpt_dir , f) for f in os.listdir(rpt_dir)]

        self.GMail(attacheds)
        #------------------------------------------------------------------

    def GMail(self , attacheds):
        #------------------------------------------------------------------
        #<<< Basic Information Of Gmail Setup  >>>
        #------------------------------------------------------------------
        GHOST = "smtp.gmail.com"
        GPORT = 587
        GMAIL_USER = ""
        GMAIL_PW = ""
        #------------------------------------------------------------------

        #------------------------------------------------------------------
        #<<< Mail - Subject >>>
        #------------------------------------------------------------------
        msg = MIMEMultipart()  #帶有附件用

        msg['Subject'] = "[指數]台灣加權指數每日資料- " + time.strftime("%Y%m%d", time.localtime())
        #------------------------------------------------------------------

        #------------------------------------------------------------------
        #<<< Mail Lists >>>
        #------------------------------------------------------------------
        TO = ["example@gmail.com" ,
              ]
        # For Debug
        #TO = ["tw.stock.lai@gmail.com" ,"smouse0220@gmail.com"]

        #Mail List Trans String
        To_Lists = ",".join(TO)
        
        FROM = GMAIL_USER

        msg['From'] = GMAIL_USER
        
        msg['To'] = To_Lists
        #------------------------------------------------------------------
        
        #------------------------------------------------------------------
        #<<< GMail Detailed Information >>>
        #------------------------------------------------------------------
        print("\n" + "*" * 25 + "\n*** GMail Information ***\n" + "*" * 25)
        print("Subjust : " , msg['Subject'])
        print("Mail Lists : " , To_Lists)
        #------------------------------------------------------------------

        #------------------------------------------------------------------
        #<<< Mail - Attached File >>>
        #------------------------------------------------------------------
        for attached in attacheds:
            att_name = os.path.basename(attached)
            
            att = MIMEText(open(attached, 'rb').read(), 'base64', 'utf-8')
            
            att["Content-Type"] = 'application/octet-stream'
            
            att.add_header('Content-Disposition', 'attachment', filename=att_name)
            
            encoders.encode_base64(att)
            
            msg.attach(att)
        #XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
        #------------------------------------------------------------------

        #------------------------------------------------------------------
        #<<< Mail - Send >>>
        #------------------------------------------------------------------
        server = smtplib.SMTP(GHOST, GPORT)
        server.ehlo()
        server.starttls()
        server.login(GMAIL_USER , GMAIL_PW)
        try:
            server.sendmail(FROM, TO , msg.as_string())
            print("\n<<<<< GMail 發送完成! >>>>>\n")
            server.quit()

        except Exception as e:
            server.quit()
            print("\n<<<<< GMail 發送失敗! >>>>>\n")
            print("Error Message : " , str(e))
        #------------------------------------------------------------------
    #..................................................................................................................

########################################################################################################
########################################################################################################

class Menu:    
    def __init__(self):
        #...........................................................................
        self.Main()
        #...........................................................................

    def Anno(self):
        #...........................................................................
        pgm_name = "tw_index" + ".exe"
        ver_info = "1.1.24"
        pgm_dir = Execute_Path()
        #rev_date = time.strftime("%Y/%m/%d" , time.localtime())
        rev_date = "2019/06/18"
        print("##############################")
        print("#" + " " *70+ "#")
        print("#	 台 灣 股 市 更 新 資 料	#")
        print("#" + " " *70+ "#")
        print("##############################")
        print("* 資料名稱 : " + pgm_name)
        print("* 程式版本 : " + ver_info)
        print("* 資料目錄 : " + pgm_dir) 
        print("* 更新日期 : " + rev_date)
        print("*" * 42)
        #...........................................................................


    def Main(self):
        #...........................................................................
        Dir_Chk()   #建立目錄
        self.Anno()
        (RCode1 , RCode2) = self.Update_Check()
        if RCode1:
            #TWII Database
            TW_DB_Create()
            TW_Index_By_Daily()
            TW_Index_By_Weekly()
            TW_Index_By_Monthly()

            #TW Futrue Database
            TX_DB_Create()
            TX_Index_By_Source()
            TX_Index_By_Daily()
            TX_Index_By_Weekly()
            TX_Index_By_Monthly()

            #TW Option Database
            TXO_DB_Create()
            TXO()

            #TW Futrue OI Database
            TXF_DB_Create()
            TXF()
            
            #Ver 1.1.18
            #TW Small Futrue OI Database
            STXF_DB_Create()
            STXF()
            
            #Ver 1.1.19
            #TXO Option
            TXOP_DB_Create()
            TXOP()
            
            #Report
            Daily_Report()
            
            #V 1.1.6A
            up_date = time.strftime("%Y-%m-%d" , time.localtime())
            db_date = Read_DB("TW" , 0)[-1][0]
            if up_date==db_date:
                GMail()
            
        else:
            while RCode2:
                print("\n <<< 手 動 更 新 程 序 >>> ")
                print("0. 離開手動更新程序")
                print("1. 台灣加權指數更新")
                print("2. 台灣期貨指數更新")
                print("3. 台灣選擇權更新")
                print("4. 台灣三大法人-大台資料")
                print("5. 台灣三大法人-小台資料")
                print("6. 台灣三大法人-選擇權資料")
                print("7. 產生日報表")
                print("8. 寄送日報表")
                ans = str(input("請輸入執行項目:"))
                if ans=="0":
                    RCode2 = False
                    
                elif ans=="1":
                    TW_DB_Create()
                    TW_Index_By_Daily()
                    TW_Index_By_Weekly()
                    TW_Index_By_Monthly()
                    print("\n<<< 完成 - 台灣加權指數更新 >>>")

                elif ans=="2":
                    TX_DB_Create()
                    TX_Index_By_Source()
                    TX_Index_By_Daily()
                    TX_Index_By_Weekly()
                    TX_Index_By_Monthly()
                    print("\n<<< 完成 - 台灣期貨指數更新 >>>")

                elif ans=="3":
                    TXO_DB_Create()
                    TXO()
                    print("\n<<< 完成 - 台灣選擇權更新 >>>")
                    
                elif ans=="4":
                    TXF_DB_Create()
                    TXF()
                    print("\n<<< 完成 - 台灣三大法人-大台資料 >>>")

                elif ans=="5":
                    STXF_DB_Create()
                    STXF()
                    print("\n<<< 完成 - 台灣三大法人-小台資料 >>>")

                elif ans=="6":
                    TXOP_DB_Create()
                    TXOP()
                    print("\n<<< 完成 - 台灣三大法人-選擇權資料 >>>")
                    
                elif ans=="7":
                    Daily_Report()
                    print("\n<<< 完成 - 產生日報表 >>>")
                    
                elif ans=="8": 
                    GMail()
                    print("\n<<< 完成 - 寄送日報表 >>>")

                else:
                    RCode2 = True
        #...........................................................................
 
    def Update_Check(self):
        #...........................................................................
        #V1.1.20
        #今（2019）年開始，針對證券及期貨市場，以及上市櫃企業有多項新制度要實施，市場投資人最關心還是周六如遇補班日，股市不再會有開市交易；此外，期貨市場的股價類期貨交易稅稅率將固定維持為10萬分之2。
        #金管會已宣布從今年起，周六補班日，證券、期貨市場一律不開市、不交割，但證券業是否上班則由各券商自行決定。
        #過往台股會因為周六補班而開市交易，形成全球獨家開市，外資法人機構操作態度消極，股市當天成交量常因而出現大幅萎縮現象，技術面遭不正常的扭曲。

        #V1.1.22
        wkday = datetime.datetime.now().weekday()
        if wkday>=5:
            print("周六/周日休市不交易")
            return(False , False)
        
        up_date = time.strftime("%Y-%m-%d" , time.localtime())

        if not os.path.exists(DB_Path()[0]):
            return(True , True)
        
        db_date = Read_DB("TW" , 0)[-1][0]
        if db_date==up_date:
            print("已經為最新資料，將進入手動模式!")
            return(False , True)
        return(True , True)
        #...........................................................................
        
########################################################################################################

#*******************************************#
# 常用功能
#*******************************************#
def Execute_Path():
    #---------------------------------------------------------------------------------------------------------------------------------
    if not sys.argv[0]==sys.executable:
        #print(".PY Execute")
        #print("Execute File : ", os.path.split(sys.argv[0])[1])
        #print("File Path : " , os.path.abspath("."))
        myPath = os.path.abspath(".")
    else:
        #print(".EXE Execute")
        #print("Execute File : ", os.path.split(sys.executable)[1])
        #print("File Path : " , os.path.split(sys.executable)[0])
        myPath = os.path.split(sys.executable)[0]
    return(myPath)
    #---------------------------------------------------------------------------------------------------------------------------------

def Dir_Chk():
    #---------------------------------------------------------------------------------------------------------------------------------
    #建立使用資料夾
    ch_dirs = ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
    
    new_ch_dirs = [os.path.join(Execute_Path() , ch_dir) for ch_dir in ch_dirs]
    
    for new_ch_dir in new_ch_dirs:
        if not os.path.exists(new_ch_dir):
            os.mkdir(new_ch_dir)
    return(new_ch_dirs)
    #---------------------------------------------------------------------------------------------------------------------------------

def DB_Path():
    #---------------------------------------------------------------------------------------------------------------------------------
    db_lists = ["twse.db", "tx.db" ,  "txo.db" , "txf.db", "stxf.db", "txop.db"]
    db_path_lits = [os.path.join(Dir_Chk()[-2] , db_list) for db_list in db_lists ]
    return(db_path_lits)
    #---------------------------------------------------------------------------------------------------------------------------------

def Move_Line(close):
    #---------------------------------------------------------------------------------------------------------------------------------
    #Ver 1.1.11
    
    Num = len(close)
    #3 , 5 , 10 , 20 , 60
    mv_data = []
    diff_data = []

    #
    t3 = 0
    t5 = 0
    t10 = 0
    t20 = 0
    t60 = 0

    d1 = 0
    d2 = 0
    d3 = 0
    d4 = 0
    for i in range(Num):
        if i >=1:
            t3 = round(sum(close[i-1:i+1])/2,2)
            
        if i>=3:
            t5 = round(sum(close[i-3:i+1])/4,2)
            d1 = round(max(t3,t5) - min(t3,t5) , 2)
            
        if i>=8:
            t10 = round(sum(close[i-8:i+1])/9,2)
            d2 = round(max(t3,t5,t10) - min(t3,t5,t10) , 2)
            
        if i>=18:
            t20 = round(sum(close[i-18:i+1])/19,2)
            d3 = round(max(t3,t5,t10,t20) - min(t3,t5,t10,t20) , 2)
            
        if i>=58:
            t60 = round(sum(close[i-58:i+1])/59,2)
            d4 = round(max(t3,t5,t10,t20,t60) - min(t3,t5,t10,t20,t60) , 2)
            
        mv_data.append([t3 , t5 , t10 , t20 , t60])
        diff_data.append([d1 , d2 , d3 , d4])
        
    return(mv_data , diff_data)
    #---------------------------------------------------------------------------------------------------------------------------------


def MA(close):
    #---------------------------------------------------------------------------------------------------------------------------------
    Num = len(close)

    #3 , 5 , 10 , 20 , 60
    ma = []

    #
    ma3 = 0
    ma5 = 0
    ma10 = 0
    ma20 = 0
    ma60 = 0
    
    for i in range(Num):
        if i >=2:ma3 = round(sum(close[i-2:i+1])/3,2)
        if i>=4:ma5 = round(sum(close[i-4:i+1])/5,2)
        if i>=9:ma10 = round(sum(close[i-9:i+1])/10,2)
        if i>=19:ma20 = round(sum(close[i-19:i+1])/20,2)
        if i>=59:ma60 = round(sum(close[i-59:i+1])/60,2)

        ma.append([ma3 , ma5 , ma10 , ma20 , ma60])
    return(ma)
    #---------------------------------------------------------------------------------------------------------------------------------

def KDJ(row_data):
    #---------------------------------------------------------------------------------------------------------------------------------
    kdj = []
    #HP = [row[1] for row in row_data]
    #LP =  [row[2] for row in row_data]
    #CP = [row[3] for row in row_data]
    hp9 = []
    lp9 = []
    for i , row in enumerate(row_data):
        hp9.append(row[1])
        lp9.append(row[2])
        if i<8:
            kdj.append([50,50,50])
        else:
            RSV = (row[-1]-min(lp9))/(max(hp9)-min(lp9)) * 100
            K = (kdj[-1][0] * 2 + RSV) / 3
            D = (kdj[-1][1] * 2 + K) / 3
            J = 3*D - 2*K
            kdj.append([K,D,J])
            del hp9[0]
            del lp9[0]
    return(kdj)
    #---------------------------------------------------------------------------------------------------------------------------------


def Daily_Report_Dir():
    #---------------------------------------------------------------------------------------------------------------------------------
    db_date = Read_DB("TXO" , 1)
        
    dates = db_date[-1][0].split("-")    #yyyy-mm-dd
        
    local_dir = Dir_Chk()[2] #["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
        
    #level1
    dir1 = os.path.join(local_dir , dates[0])
        
    if not os.path.exists(dir1):
        os.mkdir(dir1)

    #level2
    dir2 = os.path.join(dir1 , dates[0]+dates[1])
        
    if not os.path.exists(dir2):
        os.mkdir(dir2)

        #level3
    dir3 = os.path.join(dir2 , "".join(dates))
    if not os.path.exists(dir3):
        os.mkdir(dir3)
            
    return(dir3)
    #---------- DONE ----------
    #---------------------------------------------------------------------------------------------------------------------------------
   

def Close_Msg():
    #---------------------------------------------------------------------------------------------------------------------------------
    print("<<<<  台灣加權指數 & 台指期指數 網官網資料擷取完成  >>>>")
    print("*** 5秒後，本程式將自動關閉 ***")
    for i in range(5):
        print(5-i)
        time.sleep(1)
    print(0)
    #---------------------------------------------------------------------------------------------------------------------------------

def Url_Fetch(my_url , my_data):
    #.................................................................................................................................................................................................................................
    Max_Num = 5
    
    header = {
        "User-Agent" : "Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko" ,
        "Accept-Encoding" : "gzip, deflate" ,
        "Accept-Language" : "zh-TW"}

    try:
        r = requests.post(my_url, data=my_data , headers = header)
                    
    except urllib.error.URLError as e:
        if str(e)=="<urlopen error timed out>":
            time.sleep(5)
            print("Re-Fetch...")
            for i in range(Max_Num):
                try:
                    r = requests.post(my_url, data=my_data , headers = header)
                    break
                    
                except:
                    if i < Max_Num+1:
                        continue
                    else:
                        err_code = "urlerr"
                        Fetch_Error(e , my_url , err_code)
                            
        else:
            err_code = "urlerr"
            Fetch_Error(e , my_url , err_code)
                
    except urllib.error.HTTPError as e:
        err_code = "httperr"
        Fetch_Error(e , my_url , err_code)
            #Ver 1.3.3C2
        if str(e)=="HTTP Error 404: Not Found":
            pass
                
    except:
        print("Re-Fetch...")
        socket.setdefaulttimeout(10)
        time.sleep(3)
        r = requests.post(my_url, data=my_data , headers = header)
        #pass
    return(r)
    #.................................................................................................................................................................................................................................

def Fetch_Error(Err , My_Url , Err_Code):
    #---------------------------------------------------------------------------------------------------------------------------------
    ch_dirs = Dir_Chk() #ch_dirs = ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]

    #Message Descripition
    err_msg = "URL Err    : "  + str(My_Url) +"\nErr Msg    : " + str(Err) + "\nErr Reason : " + str(Err.reason) + "\n"

    #Message Box
    print("XXXXXX" * 8)
    print("URL Err    : " , My_Url)
    print("Err Msg    : " , Err)
    print("Err Reason : " , Err.reason)

    #Save Error Message File
    if Err_Code == "urlerr":
        err_file = os.path.join(ch_dirs[0] , time.strftime('%Y%m%d', time.localtime()) + "-fetch_fail(http_error).asc")
        
    else:
        print("Err Code   : " , Err.code)
        print("Err Headers: " , Err.headers)
        err_msg += "Err Code   : " + str(Err.code) + "\nErr Headers: " + str(Err.headers) + "\n"
        err_file = os.path.join(ch_dirs[0] , time.strftime('%Y%m%d', time.localtime()) + "-fetch_fail(url_error).asc")
        
    print("XXXXXX" * 8 + "\n")
    
    fp = open(err_file , "w")
    fp.write(err_msg)
    fp.close()
    #---------------------------------------------------------------------------------------------------------------------------------

def Err_Msg():
    #---------------------------------------------------------------------------------------------------------------------------------
    t,v,tb = sys.exc_info()
    err_info = traceback.extract_tb(tb)
    ch_dirs = Dir_Chk() #ch_dirs = ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
    #print(new_ch_dirs)
        
    err_msg_file =os.path.join(ch_dirs[0] , "exe_err_" + time.strftime("%Y%m%d%H%M%S", time.localtime()) + ".asc")
    f = open(err_msg_file , "w")
    f.write("Abnormal Type : " + str(t) + "\n")
    f.write("Abnormal Value : " + str(v) + "\n\n")
    f.write("<<< Abnormal Description : >>> \n")
    f.write("* File  : " + str(err_info[0][0]) + "\n")
    f.write("* Line  : " + str(err_info[0][1]) + "\n")
    f.write("* Type  : " + str(err_info[0][2]) + "\n")
    f.write("* Point : " + str(err_info[0][3]) + "\n")
    f.close()

    print("\n" * 3 + "<<< RUN-ERROR MESSAGE >>>")
    print("Abnormal Type : " + str(t))
    print("Abnormal Value : " + str(v))
    print("<<<   Abnormal Description   >>>")
    print("* File  : " + str(err_info[0][0]))
    print("* Line  : " + str(err_info[0][1]))
    print("* Type  : " + str(err_info[0][2]))
    print("* Point : " + str(err_info[0][3]))
    #---------------------------------------------------------------------------------------------------------------------------------
    
def Exe_Time(t1 , t2):
    #---------------------------------------------------------------------------------------------------------------------------------
    ch_dirs = Dir_Chk() #ch_dirs = ["Err_Msg" , "Exe_Time" ,  "Daily_Report" , "DB" , "Backup"]
    exe_time_file =os.path.join(ch_dirs[1] , "Execute_Time-" + time.strftime("%Y%m%d%H%M%S", time.localtime()) + ".asc")

    f = open(exe_time_file , "w")
    f.write("啟動時間  : " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(t1)) + "\n")
    f.write("結束時間  : " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(t2))  + "\n")
    f.write("執行時間  : " + str(round(t2-t1,3))+ "秒 \n")
    f.close()

    print("啟動時間  : " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(t1))  )
    print("結束時間  : " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(t2))  )
    print("執行時間  : " + str(round(t2-t1,3))+ "秒")
    #---------------------------------------------------------------------------------------------------------------------------------
########################################################################################################
########################################################################################################

#-----------------------------------------------------------------------------
#-----------------------------------------------------------------------------     
def MAIN():
    Menu()
#-----------------------------------------------------------------------------
#-----------------------------------------------------------------------------   


#============================================================================#
#
#============================================================================#
if __name__ == "__main__":
    time1 = time.time()
    try:
        MAIN()
    except:
        Err_Msg()
    time2 = time.time()    
    Exe_Time(time1 , time2)
#============================================================================#
